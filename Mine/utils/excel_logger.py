"""Excel logging utility for live signal monitoring."""
import os
from datetime import datetime
from typing import Optional, Dict, Any, TYPE_CHECKING
import logging

logger = logging.getLogger(__name__)

if TYPE_CHECKING:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    import openpyxl  # type: ignore[import]
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # type: ignore[import]
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl not installed. Excel logging disabled.")


# Get project root directory (relative path)
# ExcelLogger is at: /Users/kavinkumar/Mine/Mine/utils/excel_logger.py
# Go up 2 levels: Mine/utils -> Mine/
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_LOGS_DIR = os.path.join(_PROJECT_ROOT, 'logs')


class ExcelLogger:
    """Logs signal checks and trades to Excel file."""
    
    EXCEL_FILE = os.path.join(_LOGS_DIR, 'signal_logs.xlsx')
    LOGS_DIR = _LOGS_DIR
    
    def __init__(self, file_path: Optional[str] = None, username: Optional[str] = None, file_prefix: str = "signal_logs"):
        """Initialize Excel logger.
        
        Args:
            file_path: Explicit file path (takes precedence)
            username: Username for dynamic file naming (e.g., "mine" -> "mine_signal_logs.xlsx")
            file_prefix: Prefix for dynamically named files (default: "signal_logs")
        """
        self.available = EXCEL_AVAILABLE
        
        # Determine file path
        if file_path:
            # Explicit file path provided
            self.file_path = file_path
        elif username:
            # Generate dynamic filename based on username
            self.file_path = os.path.join(self.LOGS_DIR, f"{username.lower()}_{file_prefix}.xlsx")
        else:
            # Use default
            self.file_path = self.EXCEL_FILE
        
        # Ensure logs directory exists
        os.makedirs(self.LOGS_DIR, exist_ok=True)
        
        if not self.available:
            logger.warning("Excel logging disabled - install openpyxl: pip install openpyxl")
        
    def _get_or_create_workbook(self):
        """Get existing workbook or create new one."""
        if os.path.exists(self.file_path):
            return openpyxl.load_workbook(self.file_path)
        else:
            wb = openpyxl.Workbook()
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            return wb
    
    def _create_signal_checks_sheet(self, wb):
        """Create or get signal checks sheet."""
        if 'Signal Checks' not in wb.sheetnames:
            ws = wb.create_sheet('Signal Checks', 0)
            
            # Headers
            headers = [
                'Timestamp',
                'Time',
                'Market Hour',
                'CE PDH',
                'CE PDL',
                'PE PDH',
                'PE PDL',
                'CE Signal',
                'PE Signal',
                'CE Entry Price',
                'PE Entry Price',
                'CE SL',
                'PE SL',
                'CE Target',
                'PE Target',
                'Notes'
            ]
            
            ws.append(headers)
            
            # Style header
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # Set column widths
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 10
            ws.column_dimensions['G'].width = 10
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 12
            ws.column_dimensions['J'].width = 14
            ws.column_dimensions['K'].width = 14
            ws.column_dimensions['L'].width = 10
            ws.column_dimensions['M'].width = 10
            ws.column_dimensions['N'].width = 10
            ws.column_dimensions['O'].width = 10
            ws.column_dimensions['P'].width = 20
        
        return wb['Signal Checks']
    
    def _create_trades_sheet(self, wb):
        """Create or get trades sheet."""
        if 'Trades' not in wb.sheetnames:
            ws = wb.create_sheet('Trades', 1)
            
            # Headers
            headers = [
                'Timestamp',
                'Order Type',
                'Option Type',
                'Strike',
                'Entry Price',
                'Current Price',
                'Target',
                'Stop Loss',
                'PnL',
                'Status',
                'Order ID',
                'Notes'
            ]
            
            ws.append(headers)
            
            # Style header
            header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # Set column widths
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 10
            ws.column_dimensions['J'].width = 12
            ws.column_dimensions['K'].width = 15
            ws.column_dimensions['L'].width = 20
        
        return wb['Trades']
    
    def log_signal_check(self, 
                        timestamp: datetime,
                        ce_prev_high: Optional[float] = None,
                        ce_prev_low: Optional[float] = None,
                        pe_prev_high: Optional[float] = None,
                        pe_prev_low: Optional[float] = None,
                        ce_signal: bool = False,
                        pe_signal: bool = False,
                        ce_entry_price: Optional[float] = None,
                        pe_entry_price: Optional[float] = None,
                        ce_sl: Optional[float] = None,
                        pe_sl: Optional[float] = None,
                        ce_target: Optional[float] = None,
                        pe_target: Optional[float] = None,
                        notes: str = "") -> bool:
        """Log signal check to Excel.
        
        Args:
            timestamp: Timestamp of signal check
            ce_prev_high: CE previous day high
            ce_prev_low: CE previous day low
            pe_prev_high: PE previous day high
            pe_prev_low: PE previous day low
            ce_signal: Whether CE signal triggered
            pe_signal: Whether PE signal triggered
            ce_entry_price: CE entry price if signal triggered
            pe_entry_price: PE entry price if signal triggered
            ce_sl: CE stop loss
            pe_sl: PE stop loss
            ce_target: CE target
            pe_target: PE target
            notes: Additional notes
            
        Returns:
            True if logged successfully, False otherwise
        """
        if not self.available:
            return False
        
        try:
            wb = self._get_or_create_workbook()
            ws = self._create_signal_checks_sheet(wb)
            
            # Format data
            timestamp_str = timestamp.strftime('%Y-%m-%d %H:%M:%S')
            time_str = timestamp.strftime('%H:%M:%S')
            market_hour = timestamp.strftime('%H:%M')

            # Skip duplicate logs with the same timestamp + notes (prevents repeated rows)
            if ws.max_row >= 2:
                # Check recent rows to avoid repeated entries from multiple loops
                recent_start = max(2, ws.max_row - 20)
                for row_idx in range(ws.max_row, recent_start - 1, -1):
                    row_timestamp_value = ws.cell(row=row_idx, column=1).value
                    if isinstance(row_timestamp_value, datetime):
                        row_timestamp_value = row_timestamp_value.strftime('%Y-%m-%d %H:%M:%S')
                    if str(row_timestamp_value) == timestamp_str:
                        row_notes = ws.cell(row=row_idx, column=16).value
                        if (row_notes or "") == (notes or ""):
                            logger.info(f"⏭️  Duplicate signal check skipped at {timestamp_str} ({notes})")
                            return False
            
            # Prepare row
            row_data = [
                timestamp_str,
                time_str,
                market_hour,
                f"{ce_prev_high:.2f}" if ce_prev_high else "",
                f"{ce_prev_low:.2f}" if ce_prev_low else "",
                f"{pe_prev_high:.2f}" if pe_prev_high else "",
                f"{pe_prev_low:.2f}" if pe_prev_low else "",
                "✓ YES" if ce_signal else "✗ NO",
                "✓ YES" if pe_signal else "✗ NO",
                f"{ce_entry_price:.2f}" if ce_entry_price else "",
                f"{pe_entry_price:.2f}" if pe_entry_price else "",
                f"{ce_sl:.2f}" if ce_sl else "",
                f"{pe_sl:.2f}" if pe_sl else "",
                f"{ce_target:.2f}" if ce_target else "",
                f"{pe_target:.2f}" if pe_target else "",
                notes
            ]
            
            ws.append(row_data)
            
            # Apply borders and formatting
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for cell in ws[ws.max_row]:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Color code signal columns
                if cell.column == 8:  # CE Signal
                    if "YES" in str(cell.value):
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif cell.column == 9:  # PE Signal
                    if "YES" in str(cell.value):
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            
            wb.save(self.file_path)
            logger.info(f"✅ Signal check logged to Excel: {self.file_path}")
            return True
            
        except Exception as e:
            logger.error(f"❌ Error logging to Excel: {e}")
            return False
    
    def log_trade(self,
                  order_type: str,
                  option_type: str,
                  strike: int,
                  entry_price: float,
                  current_price: Optional[float] = None,
                  target: Optional[float] = None,
                  stop_loss: Optional[float] = None,
                  pnl: Optional[float] = None,
                  status: str = "OPEN",
                  order_id: Optional[str] = None,
                  notes: str = "") -> bool:
        """Log trade to Excel.
        
        Args:
            order_type: BUY or SELL
            option_type: CE or PE
            strike: Strike price
            entry_price: Entry price
            current_price: Current price
            target: Target price
            stop_loss: Stop loss level
            pnl: Profit/Loss
            status: OPEN, CLOSED, SL_HIT, TARGET_HIT, etc.
            order_id: Zerodha order ID
            notes: Additional notes
            
        Returns:
            True if logged successfully, False otherwise
        """
        if not self.available:
            return False
        
        try:
            wb = self._get_or_create_workbook()
            ws = self._create_trades_sheet(wb)
            
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Determine color based on status
            status_color = {
                'OPEN': 'FFEB9C',      # Yellow
                'BUY': 'B4C7E7',       # Light blue
                'SELL': 'C6EFCE',      # Green
                'TARGET_HIT': 'C6EFCE', # Green
                'SL_HIT': 'F8CBAD',    # Orange
                'CLOSED': 'A9D08E',    # Green
            }.get(status, 'FFFFFF')
            
            # Prepare row
            row_data = [
                timestamp,
                order_type,
                option_type,
                strike,
                f"{entry_price:.2f}",
                f"{current_price:.2f}" if current_price else "",
                f"{target:.2f}" if target else "",
                f"{stop_loss:.2f}" if stop_loss else "",
                f"{pnl:+.2f}" if pnl else "",
                status,
                order_id if order_id else "",
                notes
            ]
            
            ws.append(row_data)
            
            # Apply formatting
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for cell in ws[ws.max_row]:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Color status cell
                if cell.column == 10:  # Status
                    cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type='solid')
            
            wb.save(self.file_path)
            logger.info(f"✅ Trade logged to Excel: {order_type} {option_type} {strike}")
            return True
            
        except Exception as e:
            logger.error(f"❌ Error logging trade to Excel: {e}")
            return False
    
    def log_trailing_sl_update(self,
                               option_type: str,
                               strike: int,
                               old_sl: float,
                               new_sl: float,
                               current_price: float,
                               profit: float) -> bool:
        """Log trailing SL update to Excel.
        
        Returns:
            True if logged successfully, False otherwise
        """
        if not self.available:
            return False
        
        notes = f"Trailing SL: {old_sl:.2f} → {new_sl:.2f} (Profit: {profit:+.2f})"
        return self.log_trade(
            order_type='UPDATE',
            option_type=option_type,
            strike=strike,
            entry_price=old_sl,
            current_price=current_price,
            status='TRAILING_SL',
            notes=notes
        )
    
    def log_sl_target_check(self,
                           timestamp: datetime,
                           side: str,
                           strike: int,
                           current_price: float,
                           entry_price: float,
                           initial_sl: float,
                           target: float,
                           target_hit: bool = False,
                           trailed_sl: Optional[float] = None,
                           check_reason: str = "Monitoring") -> bool:
        """Log SL/Target check to Signal Checks sheet.
        
        Args:
            timestamp: Timestamp of check
            side: 'CE' or 'PE'
            strike: Strike price
            current_price: Current market price
            entry_price: Entry price
            initial_sl: Initial stop loss
            target: Target price
            target_hit: Whether target has been hit
            trailed_sl: Trailed SL if active (None if not trailing yet)
            check_reason: Reason for logging (e.g., "SL_HIT", "TARGET_HIT", "TRAILING", "Monitoring")
            
        Returns:
            True if logged successfully, False otherwise
        """
        if not self.available:
            return False
        
        try:
            wb = self._get_or_create_workbook()
            ws = self._create_signal_checks_sheet(wb)
            
            # Format data
            timestamp_str = timestamp.strftime('%Y-%m-%d %H:%M:%S')
            time_str = timestamp.strftime('%H:%M:%S')
            market_hour = timestamp.strftime('%H:%M')
            
            # Calculate PnL and other metrics
            pnl = current_price - entry_price
            pnl_pct = (pnl / entry_price * 100) if entry_price else 0
            
            # Determine effective SL (trailing or initial)
            effective_sl = trailed_sl if target_hit and trailed_sl is not None else initial_sl
            
            # Notes with check details
            notes = f"{side} {check_reason} | Strike: {strike} | Price: {current_price:.2f} | Entry: {entry_price:.2f} | SL: {effective_sl:.2f} | Target: {target:.2f} | PnL: {pnl:+.2f} ({pnl_pct:+.2f}%)"
            
            if target_hit and trailed_sl is not None:
                notes += f" | Trailed SL: {trailed_sl:.2f}"
            
            # Prepare row - reuse signal check structure
            row_data = [
                timestamp_str,
                time_str,
                market_hour,
                strike if side == 'CE' else "",
                "",
                strike if side == 'PE' else "",
                "",
                "YES" if side == 'CE' and check_reason != "Monitoring" else "",
                "YES" if side == 'PE' and check_reason != "Monitoring" else "",
                entry_price if side == 'CE' else "",
                entry_price if side == 'PE' else "",
                effective_sl if side == 'CE' else "",
                effective_sl if side == 'PE' else "",
                target if side == 'CE' else "",
                target if side == 'PE' else "",
                notes
            ]
            
            ws.append(row_data)
            
            # Style the row
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Color code based on reason
            if check_reason == "SL_HIT":
                cell_color = 'FF0000'  # Red
                font_color = 'FFFFFF'  # White text
            elif check_reason == "TARGET_HIT":
                cell_color = '00B050'  # Green
                font_color = 'FFFFFF'  # White text
            elif check_reason == "TRAILING":
                cell_color = 'FFC7CE'  # Light red
                font_color = '000000'  # Black text
            else:
                cell_color = 'E7E6E6'  # Light gray
                font_color = '000000'  # Black text
            
            fill = PatternFill(start_color=cell_color, end_color=cell_color, fill_type='solid')
            font = Font(color=font_color, size=10)
            
            for cell in ws[ws.max_row]:
                cell.fill = fill
                cell.font = font
                cell.border = thin_border
                if cell.column in [4, 5, 6, 7, 9, 10, 11, 12, 13, 14, 15]:  # Numeric columns
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            wb.save(self.file_path)
            return True
            
        except Exception as e:
            logger.error(f"Error logging SL/Target check: {e}", exc_info=True)
            return False


# Global instance - will be initialized with username when needed
excel_logger = None


"""
Intraday 9:20 Strategy - Live Signal Monitoring
Monitors entry signals every 30 seconds during market hours (9:15 AM - 3:30 PM IST)
"""

from datetime import datetime, timedelta, time
from typing import Dict, List, Any, Optional
import logging
import threading
import time as time_module
import os
import sys
import fcntl
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError
from .intraday_9_20 import Intraday920Strategy
from ...service.kite_order_services import KiteService
from ...service.kotak_order_services import KotakOrderService
from ...service.dhan_order_services import DhanOrderService
from ...service.fyers_order_services import FyersOrderService

# Add utils to path for excel_logger
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..', '..', '..'))
from utils.excel_logger import ExcelLogger

logger = logging.getLogger(__name__)

# ============================================================================
# CONSTANTS - Magic strings extracted to reduce duplication
# ============================================================================
# Option Types
OPTION_TYPE_CE = 'CE'
OPTION_TYPE_PE = 'PE'

# Order Status
ORDER_STATUS_SUCCESS = 'SUCCESS'
ORDER_STATUS_FAILED = 'FAILED'
ORDER_STATUS_PLACED = 'PLACED'
ORDER_STATUS_DEMO = 'DEMO'
ORDER_STATUS_ERROR = 'ERROR'

# Order Types
ORDER_TYPE_SL = 'SL_ORDER'
ORDER_TYPE_BUY = 'BUY'
ORDER_TYPE_SELL = 'SELL'

# Trade States
TRADE_STATE_OPEN = 'OPEN'
TRADE_STATE_CLOSED = 'CLOSED'
TRADE_STATE_SL_HIT = 'SL_HIT'
TRADE_STATE_TARGET_HIT = 'TARGET_HIT'
TRADE_STATE_TRAILING_SL = 'TRAILING_SL'

# Default Values
DEFAULT_NA = 'N/A'
DEFAULT_STRIKE = 0
DEFAULT_PRICE = 0.0


# ============================================================================
# HELPER FUNCTIONS - Reusable utilities to reduce code duplication
# ============================================================================
def safe_int(value: Any, default: int = DEFAULT_STRIKE) -> int:
    """Safely convert value to int with fallback to default."""
    try:
        if value:
            return int(value)
    except (ValueError, TypeError):
        pass
    return default


def safe_float(value: Any, default: float = DEFAULT_PRICE) -> float:
    """Safely convert value to float with fallback to default."""
    try:
        if isinstance(value, str):
            return float(value)
        elif value:
            return float(value)
    except (ValueError, TypeError):
        pass
    return default


def extract_option_type(side: str) -> str:
    """Extract CE or PE from side string (e.g., 'BUY CE' -> 'CE')."""
    if OPTION_TYPE_CE in str(side).upper():
        return OPTION_TYPE_CE
    elif OPTION_TYPE_PE in str(side).upper():
        return OPTION_TYPE_PE
    return DEFAULT_NA


def build_notes_string(data: Dict[str, Any], keys: List[str]) -> str:
    """Build notes string from dictionary keys that exist."""
    notes = []
    for key in keys:
        if data.get(key):
            notes.append(f"{key.capitalize()}: {data[key]}")
    return " | ".join(notes) if notes else ""


class _NoOpExcelLogger:
    """Null object pattern - provides all ExcelLogger methods but does nothing."""
    file_path = None
    
    def log_trade(self, **kwargs):
        pass
    
    def log_signal_check(self, **kwargs):
        pass
    
    def log_sl_target_check(self, **kwargs):
        pass
    
    def log_trailing_sl_update(self, **kwargs):
        pass


# Initialize excel logger - will be set per user when monitoring starts
# Use NoOpLogger as default to avoid None checks everywhere
excel_logger: ExcelLogger = _NoOpExcelLogger()  # type: ignore


def init_excel_logger(username: Optional[str] = None, file_prefix: str = "signal_logs") -> None:
    """Initialize the excel logger for Intraday 9:20 monitoring.
    
    Args:
        username: Username for Excel logger file naming (optional)
        file_prefix: Prefix for the Excel file (default: "signal_logs")
    """
    global excel_logger
    excel_logger = ExcelLogger(username=username, file_prefix=file_prefix)
    logger.info(f"Excel logger initialized: {excel_logger.file_path}")


def log_order_placement(order_data: Dict[str, Any]) -> None:
    """
    Log order placement details to Excel file.
    
    Args:
        order_data: Dictionary containing order details
    """
    try:
        side = order_data.get('side', DEFAULT_NA)
        status = order_data.get('status', DEFAULT_NA).upper()
        
        # Map status to Excel logger status
        if ORDER_STATUS_SUCCESS in status:
            excel_status = side
        elif ORDER_STATUS_FAILED in status or ORDER_STATUS_ERROR in status:
            excel_status = ORDER_STATUS_FAILED
        elif ORDER_STATUS_DEMO in order_data.get('mode', ''):
            excel_status = f"{side}_{ORDER_STATUS_DEMO}"
        else:
            excel_status = status
        
        # Prepare notes from various fields
        notes_str = build_notes_string(order_data, ['mode', 'error', 'details', 'order_type'])
        option_type = extract_option_type(side)
        strike = safe_int(order_data.get('strike'))
        entry_price = safe_float(order_data.get('entry_price'))
        
        # Handle target and SL safely
        target_price = None
        target_val = order_data.get('target')
        if target_val not in [DEFAULT_NA, None, '']:
            target_price = safe_float(target_val)
        
        sl_price = None
        sl_val = order_data.get('sl')
        if sl_val not in [DEFAULT_NA, None, '']:
            sl_price = safe_float(sl_val)
        
        # Log to Excel
        excel_logger.log_trade(
            order_type=side,
            option_type=option_type,
            strike=strike,
            entry_price=entry_price,
            current_price=entry_price,
            target=target_price,
            stop_loss=sl_price,
            pnl=None,
            status=excel_status,
            order_id=order_data.get('order_id'),
            notes=notes_str
        )
        logger.info(f"✅ Order placed: {side}")
        
    except Exception as e:
        logger.error(f"Failed to write order placement log: {e}", exc_info=True)


class Intraday920LiveSignal:
    """
    Live signal monitoring for Intraday 9:20 strategy.
    
    Features:
    - Monitors every 30 seconds
    - Only during market hours (9:15 AM - 3:30 PM IST)
    - Tracks entry signals in real-time
    - Maintains trade state across monitoring intervals
    """
    
    # Market hours (IST)
    MARKET_OPEN = time(9, 15, 0)      # 9:15 AM
    MARKET_CLOSE = time(15, 20, 0)    # 3:20 PM (market closes at 3:30 but last candle is 3:20)
    MONITORING_INTERVAL = 3  # seconds
    
    def __init__(self, kite_instance, symbol: str = 'NIFTY', live_trading: bool = True, username: Optional[str] = None):
        """
        Initialize live signal monitor.
        
        Args:
            kite_instance: KiteConnect instance
            symbol: Trading symbol (NIFTY, BANKNIFTY, FINNIFTY)
            live_trading: Whether to place real orders (default: True for live trading)
            username: Username for Excel logger file naming (optional)
        """
        self.kite = kite_instance
        self.symbol = symbol
        self.username = username
        self.strategy = Intraday920Strategy(kite_instance)
        self.kite_service = KiteService(kite_instance=kite_instance)
        
        # Initialize Multi-Broker Support
        self.extra_brokers = {}  # Dictionary to hold active broker services
        self._init_extra_brokers()
        
        # Initialize Excel logger for this user
        init_excel_logger(username=username, file_prefix="signal_logs")
        
        # Monitoring state
        self.is_monitoring = False
        self.monitor_thread = None
        self.is_entry_monitoring = False
        self.entry_monitor_thread = None
        self.is_sl_monitoring = False
        self.sl_monitor_thread = None
        self.active_trades = {}  # Track active trades: {side: {entry_info, order_id, target_hit, trailed_sl}}
        self.active_trades_lock = threading.RLock()  # Reentrant lock - allows nested acquisition (e.g., close_trade called from within check_sl_target)
        self.today_signals = []  # All signals generated today
        self.daily_entries = {}  # Track entries per side per day: {side: entry_date} to prevent multiple entries per day
        self.daily_entries_lock = threading.Lock()  # Thread safety for daily_entries
        self.last_entry_check_time = None  # Track last entry signal check to prevent duplicates
        self.last_sl_target_check_time = None  # Track last SL/Target check time (for time-based intervals)
        self.market_close_processed = False  # Flag to ensure market close only happens once per day
        
        # Trading configuration
        self.live_trading = live_trading  # False=demo mode, True=live orders
        self.risk_reward_ratio = '1:2-trail'  # Use 1:2 with trailing SL
        
        # Configuration constants
        self.LIVE_DATA_FETCH_TIMEOUT = 10  # seconds
        self.PRICE_FETCH_BATCH_SIZE = 10  # max tokens per quote call
        
        # Process locking
        self.lock_file = f"/tmp/intraday_9_20_{username if username else 'default'}.lock"
        self.lock_fd = None

        
        
        logger.info(f"Intraday 9:20 Live Signal Monitor initialized for {symbol} (live_trading={live_trading}, ratio={self.risk_reward_ratio}) [Instance ID: {id(self)}]")
    
    def _init_extra_brokers(self):
        """Initialize additional broker services if credentials are present in env."""
        # 1. Kotak Neo
        if os.getenv("KOTAK_ACCESS_TOKEN") and os.getenv("KOTAK_MOBILE_NUMBER"):
            try:
                self.extra_brokers['KOTAK'] = KotakOrderService()
                logger.info("✅ Kotak Neo Service Initialized")
            except Exception as e:
                logger.error(f"❌ Failed to init Kotak Neo: {e}")
        
        # 2. Dhan
        if os.getenv("DHAN_ACCESS_TOKEN") and os.getenv("DHAN_CLIENT_ID"):
            try:
                self.extra_brokers['DHAN'] = DhanOrderService()
                logger.info("✅ Dhan Service Initialized")
            except Exception as e:
                logger.error(f"❌ Failed to init Dhan: {e}")

        # 3. Fyers
        if os.getenv("FYERS_APP_ID") and (os.getenv("FYERS_ACCESS_TOKEN") or os.getenv("FYERS_SECRET_KEY")):
            try:
                self.extra_brokers['FYERS'] = FyersOrderService()
                logger.info("✅ Fyers Service Initialized")
            except Exception as e:
                logger.error(f"❌ Failed to init Fyers: {e}")
                
        if self.extra_brokers:
            logger.info(f"Active Extra Brokers: {list(self.extra_brokers.keys())}")

    def _acquire_process_lock(self) -> bool:
        """
        Acquire file lock to ensure only one monitor instance runs per user.
        Prevents multiple processes (e.g., Gunicorn workers) from running duplicate monitors.
        """
        try:
            self.lock_fd = open(self.lock_file, 'w')
            # Try to acquire an exclusive lock - non-blocking
            fcntl.flock(self.lock_fd, fcntl.LOCK_EX | fcntl.LOCK_NB)
            # Write PID to lock file
            self.lock_fd.write(str(os.getpid()))
            self.lock_fd.flush()
            logger.info(f"🔒 Process lock acquired: {self.lock_file} (PID: {os.getpid()})")
            return True
        except (IOError, OSError):
            logger.warning(f"⚠️ Could not acquire process lock for {self.username}. Another instance is running.")
            if self.lock_fd:
                try:
                    self.lock_fd.close()
                except:
                    pass
                self.lock_fd = None
            return False
            
    def _release_process_lock(self) -> None:
        """Release the process lock."""
        if self.lock_fd:
            try:
                fcntl.flock(self.lock_fd, fcntl.LOCK_UN)
                self.lock_fd.close()
                logger.info(f"🔓 Process lock released: {self.lock_file}")
            except Exception as e:
                logger.error(f"Error releasing process lock: {e}")
            finally:
                self.lock_fd = None

    
    def is_market_hours(self, check_time: Optional[datetime] = None) -> bool:
        """
        Check if current time is within market hours (9:15 AM - 3:30 PM IST).
        
        Args:
            check_time: Optional datetime to check. Defaults to now.
            
        Returns:
            True if within market hours, False otherwise
        """
        if check_time is None:
            check_time = datetime.now()
        
        current_time = check_time.time()
        
        # Market hours: 9:15 AM to 3:30 PM
        return self.MARKET_OPEN <= current_time <= self.MARKET_CLOSE
    
    def is_market_day(self, check_date: Optional[datetime] = None) -> bool:
        """
        Check if date is a market trading day (Mon-Fri, not weekend).
        
        Args:
            check_date: Optional datetime to check. Defaults to today.
            
        Returns:
            True if trading day, False otherwise
        """
        if check_date is None:
            check_date = datetime.now()
        
        # Monday=0, Sunday=6
        return check_date.weekday() < 5
    
    def should_check_sl_target_now(self) -> bool:
        """
        Determine if we should check for SL/TARGET based on 3-second time intervals.
        Uses time-delta approach instead of second modulo for better accuracy and reliability.
        Prevents race conditions from API delays or loop processing time.
        
        Returns:
            True if 3+ seconds have elapsed since last check, False otherwise
        """
        now = datetime.now()
        
        # First check ever - do it now
        if self.last_sl_target_check_time is None:
            self.last_sl_target_check_time = now
            return True
        
        # Check if 3+ seconds have elapsed
        elapsed = (now - self.last_sl_target_check_time).total_seconds()
        if elapsed >= self.MONITORING_INTERVAL:  # 3 seconds
            self.last_sl_target_check_time = now
            return True
        
        return False
    
    def has_entered_today(self, side: str, strike: int) -> bool:
        """
        Check if entry has already been made for this side+strike combination today.
        Prevents multiple entries for the same strike+side per day.
        e.g., CE_24600, PE_24600 are tracked separately
        
        Checks both:
        1. In-memory daily_entries (current session)
        2. Excel Trade sheet (persistent across restarts)
        
        Args:
            side: 'CE' or 'PE'
            strike: Strike price (e.g., 24600)
            
        Returns:
            True if already entered today for this strike+side, False otherwise
        """
        today = datetime.now().date()
        entry_key = f"{side}_{strike}"  # e.g., 'CE_24600'
        
        # Check in-memory tracking first
        with self.daily_entries_lock:
            if entry_key in self.daily_entries:
                entry_date = self.daily_entries[entry_key]
                if entry_date == today:
                    logger.info(f"⛔ {side} {strike} found in daily_entries (in-memory tracking)")
                    return True
        
        # Check Excel Trade sheet for persistent tracking (survives app restarts)
        try:
            from openpyxl import load_workbook
            import os
            
            # Only check Excel if logger is initialized
            if not excel_logger or not excel_logger.file_path:
                return False
            
            if os.path.exists(excel_logger.file_path):
                wb = load_workbook(excel_logger.file_path)
                if 'Trades' in wb.sheetnames:
                    ws = wb['Trades']
                    today_str = today.strftime('%Y-%m-%d')
                    
                    # Search for entries with matching strike, side, and today's date
                    for row_idx in range(2, ws.max_row + 1):
                        row_date = ws.cell(row=row_idx, column=1).value  # Timestamp column
                        row_order_type = ws.cell(row=row_idx, column=2).value  # Order Type
                        row_option_type = ws.cell(row=row_idx, column=3).value  # Option Type
                        row_strike = ws.cell(row=row_idx, column=4).value  # Strike column
                        
                        # Check if this row matches our criteria
                        if row_date and row_option_type and row_strike:
                            row_date_str = str(row_date)[:10] if row_date else ""
                            
                            if (row_date_str == today_str and 
                                row_option_type == side and 
                                row_strike == strike and
                                row_order_type == 'BUY'):
                                logger.info(f"⛔ {side} {strike} found in Excel Trade sheet for today ({today_str})")
                                # Mark in memory for faster lookup next time
                                with self.daily_entries_lock:
                                    self.daily_entries[entry_key] = today
                                return True
        except Exception as e:
            logger.warning(f"Could not check Excel Trade sheet for duplicates: {e}")
        
        return False
    
    def mark_entry_today(self, side: str, strike: int) -> None:
        """
        Mark that an entry has been made for this side+strike combination today.
        
        Args:
            side: 'CE' or 'PE'
            strike: Strike price (e.g., 24600)
        """
        today = datetime.now().date()
        entry_key = f"{side}_{strike}"  # e.g., 'CE_24600'
        with self.daily_entries_lock:
            self.daily_entries[entry_key] = today
            logger.info(f"📋 {side} {strike} entry marked for today ({today})")
    
    def reset_daily_entries(self) -> None:
        """
        Reset daily entries tracking (call at market close or new day).
        """
        with self.daily_entries_lock:
            self.daily_entries = {}
            self.market_close_processed = False  # Reset market close flag for next day
            logger.info("🔄 Daily entries tracking reset & market close flag reset")
    
    def get_first_entry_time_today(self, side: str, strike: int) -> Optional[float]:
        """
        Get the FIRST entry time for a strike+side combination today from Excel.
        Used to ensure only the FIRST entry is used for that strike, matching backtest logic.
        
        Args:
            side: 'CE' or 'PE'
            strike: Strike price
            
        Returns:
            Entry timestamp (Unix) if found, None if not found or no entries today
        """
        today = datetime.now().date()
        
        try:
            from openpyxl import load_workbook
            import os
            
            if not excel_logger or not excel_logger.file_path:
                return None
            
            if os.path.exists(excel_logger.file_path):
                wb = load_workbook(excel_logger.file_path)
                if 'Trades' in wb.sheetnames:
                    ws = wb['Trades']
                    today_str = today.strftime('%Y-%m-%d')
                    
                    # Find ALL BUY entries for this strike+side today, return the FIRST one
                    first_entry_time = None
                    
                    for row_idx in range(2, ws.max_row + 1):
                        row_date = ws.cell(row=row_idx, column=1).value  # Timestamp
                        row_order_type = ws.cell(row=row_idx, column=2).value  # Order Type
                        row_option_type = ws.cell(row=row_idx, column=3).value  # Option Type
                        row_strike = ws.cell(row=row_idx, column=4).value  # Strike
                        
                        if row_date and row_option_type and row_strike:
                            row_date_str = str(row_date)[:10] if row_date else ""
                            
                            if (row_date_str == today_str and 
                                row_option_type == side and 
                                row_strike == strike and
                                row_order_type == 'BUY'):
                                
                                # Convert row_date to timestamp for comparison
                                entry_timestamp = None
                                if isinstance(row_date, str):
                                    try:
                                        entry_dt = datetime.strptime(row_date, '%Y-%m-%d %H:%M:%S')
                                        entry_timestamp = entry_dt.timestamp()
                                    except (ValueError, TypeError):
                                        entry_timestamp = None
                                elif isinstance(row_date, datetime):
                                    entry_timestamp = row_date.timestamp()
                                
                                # Keep track of the earliest (first) entry
                                if entry_timestamp:
                                    if first_entry_time is None or entry_timestamp < first_entry_time:
                                        first_entry_time = entry_timestamp
                    
                    if first_entry_time:
                        logger.info(f"✅ Found first entry for {side} {strike} today: {first_entry_time}")
                    
                    return first_entry_time
        except Exception as e:
            logger.warning(f"Could not check first entry time from Excel: {e}")
        
        return None
    
    def should_check_entry_signal(self) -> bool:
        """
        Determine if we should check for ENTRY signals now.
        Returns True only at 5-minute marks (9:15, 9:20, 9:25, ..., 3:15, 3:20).
        
        Uses last_entry_check_time to ensure each 5-minute interval is checked exactly once,
        preventing missed checks after hour boundaries or during processing delays.
        
        Entry signals checked only at 5-minute intervals.
        SL/Target checks happen every 30 seconds.
        
        Returns:
            True if we're at a 5-minute mark AND haven't checked this interval yet
        """
        now = datetime.now()
        
        # Check if we're at a 5-minute mark (minute divisible by 5)
        if now.minute % 5 != 0:
            return False
        
        # Calculate current 5-minute interval (e.g., 14:00, 14:05, 14:10)
        current_interval = now.replace(second=0, microsecond=0)
        
        # If we haven't checked yet, or if this is a new 5-minute interval
        if self.last_entry_check_time is None or current_interval > self.last_entry_check_time:
            self.last_entry_check_time = current_interval
            return True
        
        return False
    
    async def fetch_live_data(self) -> Dict[str, Any]:
        """
        Fetch live market data for CE/PE strikes.
        
        Returns:
            Dictionary with current strike data
        """
        try:
            # Get current price and first 5-min data
            intraday_data = self.strategy.get_intraday_920_data(self.symbol)
            
            if not intraday_data.get('success'):
                logger.warning(f"Failed to fetch intraday data: {intraday_data.get('error')}")
                return {'success': False}
            
            return intraday_data
            
        except Exception as e:
            logger.error(f"Error fetching live data for {self.symbol}: {str(e)}")
            return {'success': False, 'error': str(e)}

    def _fetch_live_data_with_refresh(self) -> Dict[str, Any]:
        """Fetch live data with retry logic for connection resilience.
        
        Strategy:
        1. Try to fetch live data (get current price and 5-minute candles)
        2. If token invalid, refresh token and retry
        3. If connection error, retry up to 3 times with exponential backoff
        """
        import time as time_module
        
        max_retries = 3
        retry_delay = 0.5  # Start with 500ms
        last_error = None
        
        for attempt in range(max_retries):
            try:
                fetch_start = time_module.time()
                logger.info(f"[Data Fetch] Attempt {attempt + 1}/{max_retries}: Fetching live data for {self.symbol}...")
                live_data = self.strategy.get_intraday_920_data(self.symbol)
                fetch_time = time_module.time() - fetch_start
                
                if live_data.get('success'):
                    logger.info(f"[Data Fetch] ✅ Success in {fetch_time:.2f}s")
                    return live_data
                
                # Check if error is token-related
                error_msg = str(live_data.get('error', ''))
                logger.warning(f"[Data Fetch] Failed in {fetch_time:.2f}s: {error_msg}")
                
                if 'Incorrect `api_key` or `access_token`' in error_msg:
                    logger.warning(f"[Live Data Fetch] Access token invalid (attempt {attempt + 1}/{max_retries}). Attempting refresh...")
                    try:
                        refresh_ok = self.strategy._refresh_kite_access_token()  # type: ignore[attr-defined]
                        if refresh_ok:
                            # Retry after token refresh
                            time_module.sleep(retry_delay)
                            continue
                    except Exception as e:
                        logger.warning(f"Token refresh failed: {e}")
                
                # Check if error is retriable (connection reset, timeout, etc)
                is_retriable = any(keyword in error_msg.lower() for keyword in [
                    'connection reset', 'connection aborted', 'connection refused',
                    'timeout', 'gateway', '504', '503'
                ])
                
                if is_retriable and attempt < max_retries - 1:
                    logger.warning(f"[Live Data Fetch] Retriable error (attempt {attempt + 1}/{max_retries}): {error_msg}")
                    last_error = error_msg
                    # Exponential backoff
                    time_module.sleep(retry_delay * (2 ** attempt))
                    continue
                else:
                    # Non-retriable error
                    logger.error(f"[Live Data Fetch] Failed to fetch live data: {error_msg}")
                    return live_data
                    
            except Exception as e:
                logger.warning(f"[Live Data Fetch] Exception (attempt {attempt + 1}/{max_retries}): {e}")
                last_error = str(e)
                if attempt < max_retries - 1:
                    time_module.sleep(retry_delay * (2 ** attempt))
                    continue
                else:
                    return {
                        'success': False,
                        'error': f'Failed after {max_retries} retries: {last_error}'
                    }
        
        # All retries exhausted
        return {
            'success': False,
            'error': f'Failed to fetch live data after {max_retries} retries. Last error: {last_error}'
        }


    def _fetch_live_data_with_timeout(self, timeout_seconds: int = 45) -> Dict[str, Any]:
        """Fetch live data with a timeout to avoid blocking the 5-min log cycle.
        
        Uses ThreadPoolExecutor to run fetch in a separate thread with timeout protection.
        Increased timeout to 45 seconds to allow for candle fetching with retries.
        
        Args:
            timeout_seconds: Maximum time to wait for live data fetch (default: 45s)
            
        Returns:
            Live data dictionary or error response
        """
        with ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(self._fetch_live_data_with_refresh)
            try:
                return future.result(timeout=timeout_seconds)
            except FuturesTimeoutError:
                logger.error(f"[Live Data Fetch] Timeout after {timeout_seconds}s - candle fetching may be slow")
                return {
                    'success': False,
                    'error': f'Live data fetch timed out after {timeout_seconds}s - candle fetching took too long, retrying next cycle'
                }
            except Exception as e:
                logger.error(f"[Live Data Fetch] Unexpected error: {e}")
                return {
                    'success': False,
                    'error': f'Failed to fetch live data: {str(e)}'
                }
    
    def check_entry_signals_parallel(self, high_strike: Dict[str, Any], low_strike: Dict[str, Any]) -> tuple:
        """
        Check entry signals for both high and low strikes in PARALLEL for efficiency.
        
        Parallel execution reduces wait time when checking multiple strikes.
        
        Args:
            high_strike: High strike data
            low_strike: Low strike data
            
        Returns:
            Tuple of (high_signals, low_signals)
        """
        with ThreadPoolExecutor(max_workers=2) as executor:
            high_future = executor.submit(
                self.check_entry_signal_live,
                int(high_strike.get('ce_token', 0)),
                int(high_strike.get('pe_token', 0)),
                float(high_strike.get('ce_high', 0)),
                float(high_strike.get('pe_high', 0))
            )
            
            low_future = executor.submit(
                self.check_entry_signal_live,
                int(low_strike.get('ce_token', 0)),
                int(low_strike.get('pe_token', 0)),
                float(low_strike.get('ce_high', 0)),
                float(low_strike.get('pe_high', 0))
            )
            
            try:
                high_signals = high_future.result(timeout=self.LIVE_DATA_FETCH_TIMEOUT)
                low_signals = low_future.result(timeout=self.LIVE_DATA_FETCH_TIMEOUT)
            except FuturesTimeoutError:
                logger.warning("Signal check timed out")
                high_signals = {'success': False, 'error': 'Timeout'}
                low_signals = {'success': False, 'error': 'Timeout'}
            
            return high_signals, low_signals
    
    def _process_strike_signals(self, signals: Dict[str, Any], strike_data: Dict[str, Any], 
                               strike_name: str) -> Dict[str, Any]:
        """
        Process entry signals from a single strike (DRY pattern).
        
        Args:
            signals: Signal data from check_entry_signal_live()
            strike_data: Strike data for order placement
            strike_name: Name for logging ("HIGH STRIKE" or "LOW STRIKE")
            
        Returns:
            Dictionary with signal processing results
        """
        result = {
            'has_ce_signal': False,
            'has_pe_signal': False,
            'ce_entry_price': None,
            'ce_sl': None,
            'ce_target': None,
            'pe_entry_price': None,
            'pe_sl': None,
            'pe_target': None,
            'ce_high_val': strike_data.get('ce_high'),
            'pe_high_val': strike_data.get('pe_high')
        }
        
        if not signals.get('success'):
            return result
        
        ce_sig = signals.get('ce_signal', {})
        pe_sig = signals.get('pe_signal', {})
        
        logger.info(f"{strike_name} Signals - CE has_signal: {ce_sig.get('has_signal')}, PE has_signal: {pe_sig.get('has_signal')}")
        
        if ce_sig.get('has_signal'):
            result['has_ce_signal'] = True
            result['ce_entry_price'] = ce_sig.get('entry_price')
            result['ce_sl'] = ce_sig.get('sl')
            result['ce_target'] = ce_sig.get('target')
            logger.info(f"📊 {strike_name} CE SIGNAL: Entry {ce_sig.get('entry_price')}, SL {ce_sig.get('sl')}, Target {ce_sig.get('target')}")
        else:
            logger.info(f"CE No Signal Reason: {ce_sig.get('reason', 'Unknown')}")
        
        if pe_sig.get('has_signal'):
            result['has_pe_signal'] = True
            result['pe_entry_price'] = pe_sig.get('entry_price')
            result['pe_sl'] = pe_sig.get('sl')
            result['pe_target'] = pe_sig.get('target')
            logger.info(f"📊 {strike_name} PE SIGNAL: Entry {pe_sig.get('entry_price')}, SL {pe_sig.get('sl')}, Target {pe_sig.get('target')}")
        else:
            logger.info(f"PE No Signal Reason: {pe_sig.get('reason', 'Unknown')}")
        
        # Update trades if there are signals
        if result['has_ce_signal'] or result['has_pe_signal']:
            self.update_active_trades(signals, strike_data=strike_data)
            self.log_signal(signals)
        
        return result
    
    def update_active_trades(self, signals: Dict[str, Any], strike_data: Optional[Dict[str, Any]] = None) -> None:
        """
        Update active trade state with new signals from strategy.check_entry_signal().
        Places buy orders when entry signals are detected.
        
        IMPORTANT: Allows multiple sequential entries per side (e.g., 10:15 CE entry, exit at 11:45, then 11:00 CE entry).
        New entries will replace the existing trade on that side only if the previous trade has been closed.
        
        Uses signal data returned by the strategy which includes:
        - has_signal: Whether entry conditions were met
        - entry_price: Entry price at signal
        - sl: Stop loss calculated by strategy.calculate_sl_for_entry()
        - target: Target calculated by strategy.calculate_sl_for_entry()
        
        Thread-safe access to active_trades dictionary.
        
        Args:
            signals: Entry signals from check_entry_signal_live()
            strike_data: Optional strike data for order placement {ce_token, pe_token, ce_high, pe_high}
        """
        ce_signal = signals.get('ce_signal', {})
        pe_signal = signals.get('pe_signal', {})
        
        with self.active_trades_lock:
            # Track CE entry and place buy order
            # Allow new CE entry ONLY if:
            # 1. Signal exists AND (no CE trade exists OR existing CE trade is already closed)
            # 2. AND no entry has been made for this CE STRIKE today (prevent multiple entries per strike per day)
            # LOGIC MATCHES BACKTEST: Only the FIRST entry for a strike is accepted, others are rejected
            if ce_signal.get('has_signal'):
                order_id = None
                ce_strike = None
                if strike_data:
                    # Get actual CE strike price (not the candle high)
                    ce_strike_val = strike_data.get('ce_strike')
                    if ce_strike_val:
                        ce_strike = int(ce_strike_val)
                    else:
                        ce_high_val = strike_data.get('ce_high')
                        if ce_high_val:
                            ce_strike = int(ce_high_val)
                
                # Check if already entered today for this CE strike
                # This prevents duplicate entries matching backtest "only first entry" logic
                if ce_strike and self.has_entered_today('CE', ce_strike):
                    logger.info(f"⛔ CE {ce_strike} entry already made today - rejecting subsequent entry signals (matches backtest logic: only first entry per strike)")
                elif 'CE' not in self.active_trades or self.active_trades['CE'].get('status') == 'CLOSED':
                    # Place order FIRST, only mark as entered if successful
                    order_id = None
                    if strike_data and ce_strike:
                        order_id = self.place_buy_order(
                            side='CE',
                            token=strike_data.get('ce_token'),  # type: ignore
                            strike=ce_strike,
                            entry_price=ce_signal.get('entry_price')
                        )
                        
                        # Mark entry ONLY if order was successfully placed
                        if order_id:
                            self.mark_entry_today('CE', ce_strike)
                    
                    # Place SL order on broker
                    sl_order_id = None
                    if self.live_trading and strike_data and ce_strike:
                        try:
                            # Get lot size for this symbol (same as entry order quantity)
                            entry_quantity = self.kite_service.get_lot_size(self.symbol)
                            
                            # Get option trading symbol for SL order
                            option_symbol = self._get_option_symbol(self.symbol, ce_strike, 'CE')
                            if option_symbol:
                                logger.info(f"Placing CE SL order with symbol: {option_symbol}, quantity: {entry_quantity}")
                                sl_result = self.kite_service.place_stoploss_order(
                                    tradingsymbol=option_symbol,
                                    trigger_price=ce_signal.get('sl'),
                                    quantity=entry_quantity,
                                    product='NRML'
                                )
                                
                                # --- Multi-Broker SL Order ---
                                extra_sl_ids = {}
                                if self.extra_brokers and self.live_trading:
                                    logger.info(f"⚡ Placing COPY SL orders on {len(self.extra_brokers)} extra brokers...")
                                    
                                    with ThreadPoolExecutor(max_workers=3) as executor:
                                        futures = {
                                            executor.submit(
                                                self._place_extra_broker_sl_order, 
                                                name, service, 'CE', ce_strike, ce_signal.get('sl')
                                            ): name for name, service in self.extra_brokers.items()
                                        }
                                        
                                        for future in futures:
                                            try:
                                                result = future.result(timeout=5)
                                                if result:
                                                    b_name, b_order_id = result
                                                    extra_sl_ids[b_name] = b_order_id
                                            except Exception as e:
                                                logger.error(f"SL placement failed for {futures[future]}: {e}")
                                # -----------------------------

                                if sl_result['success']:
                                    sl_order_id = sl_result['order_id']
                                    logger.info(f"✅ CE SL order placed: {option_symbol} @ {ce_signal.get('sl'):.2f} | SL Order ID: {sl_order_id}")
                                    # Log SL order to Trade sheet
                                    excel_logger.log_trade(
                                        order_type='SL_ORDER',
                                        option_type='CE',
                                        strike=ce_strike,
                                        entry_price=ce_signal.get('sl'),
                                        target=ce_signal.get('target'),
                                        stop_loss=ce_signal.get('sl'),
                                        status='PLACED',
                                        order_id=sl_order_id,
                                        notes=f'SL Order Placed | Trigger: {ce_signal.get("sl"):.2f}'
                                    )
                                else:
                                    logger.error(f"❌ Failed to place CE SL order: {sl_result.get('error')}")
                                    # Log failure to Trade sheet
                                    excel_logger.log_trade(
                                        order_type='SL_ORDER',
                                        option_type='CE',
                                        strike=ce_strike,
                                        entry_price=ce_signal.get('sl'),
                                        target=ce_signal.get('target'),
                                        stop_loss=ce_signal.get('sl'),
                                        status='FAILED',
                                        notes=f'SL Order Failed: {sl_result.get("error")}'
                                    )
                            else:
                                logger.warning(f"Could not find option symbol for CE {ce_strike} - SL order not placed. Will monitor internally.")
                        except Exception as e:
                            logger.error(f"Error placing CE SL order: {e}", exc_info=True)
                    
                    self.active_trades['CE'] = {
                        'entry_price': ce_signal.get('entry_price'),
                        'entry_high': ce_signal.get('entry_high'),  # Reference high used for entry
                        'sl': ce_signal.get('sl'),
                        'target': ce_signal.get('target'),
                        'entry_time': datetime.now().isoformat(),
                        'order_id': order_id,
                        'sl_order_id': sl_order_id,  # Track SL order ID for modifications
                        'extra_sl_ids': extra_sl_ids, # Track SL IDs for other brokers
                        'token': strike_data.get('ce_token') if strike_data else None,  # type: ignore
                        'strike': ce_strike if strike_data else None,
                        'status': 'OPEN',
                        # Trailing SL state
                        'target_hit': False,  # Track if target was hit
                        'trailed_sl': ce_signal.get('sl'),  # Current trailed SL (starts at initial SL)
                        'sl_distance': ce_signal.get('entry_price') - ce_signal.get('sl')  # Distance between entry and SL
                    }
                    # Entry is already marked above to prevent race conditions
                    logger.info(f"🟢 CE {ce_strike} trade opened at {ce_signal.get('entry_price')} (Entry High: {ce_signal.get('entry_high')}, SL: {ce_signal.get('sl')}, Target: {ce_signal.get('target')}) | Order ID: {order_id if order_id else 'N/A'} | SL Order ID: {sl_order_id if sl_order_id else 'N/A'}")
                else:
                    logger.info(f"⏭️  CE signal detected but trade already OPEN - skipping to allow sequential entries (close existing trade first)")
            
            # Track PE entry and place buy order
            # Allow new PE entry ONLY if:
            # 1. Signal exists AND (no PE trade exists OR existing PE trade is already closed)
            # 2. AND no entry has been made for this PE STRIKE today (prevent multiple entries per strike per day)
            # LOGIC MATCHES BACKTEST: Only the FIRST entry for a strike is accepted, others are rejected
            if pe_signal.get('has_signal'):
                order_id = None
                pe_strike = None
                if strike_data:
                    # Get actual PE strike price (not the candle high)
                    pe_strike_val = strike_data.get('pe_strike')
                    if pe_strike_val:
                        pe_strike = int(pe_strike_val)
                    else:
                        pe_high_val = strike_data.get('pe_high')
                        if pe_high_val:
                            pe_strike = int(pe_high_val)
                
                # Check if already entered today for this PE strike
                # This prevents duplicate entries matching backtest "only first entry" logic
                if pe_strike and self.has_entered_today('PE', pe_strike):
                    logger.info(f"⛔ PE {pe_strike} entry already made today - rejecting subsequent entry signals (matches backtest logic: only first entry per strike)")
                elif 'PE' not in self.active_trades or self.active_trades['PE'].get('status') == 'CLOSED':
                    # Place order FIRST, only mark as entered if successful
                    order_id = None
                    if strike_data and pe_strike:
                        order_id = self.place_buy_order(
                            side='PE',
                            token=strike_data.get('pe_token'),  # type: ignore
                            strike=pe_strike,
                            entry_price=pe_signal.get('entry_price')
                        )
                        
                        # Mark entry ONLY if order was successfully placed
                        if order_id:
                            self.mark_entry_today('PE', pe_strike)
                    
                    # Place SL order on broker
                    sl_order_id = None
                    if self.live_trading and strike_data and pe_strike:
                        try:
                            # Get lot size for this symbol (same as entry order quantity)
                            entry_quantity = self.kite_service.get_lot_size(self.symbol)
                            
                            # Get option trading symbol for SL order
                            option_symbol = self._get_option_symbol(self.symbol, pe_strike, 'PE')
                            if option_symbol:
                                logger.info(f"Placing PE SL order with symbol: {option_symbol}, quantity: {entry_quantity}")
                                sl_result = self.kite_service.place_stoploss_order(
                                    tradingsymbol=option_symbol,
                                    trigger_price=pe_signal.get('sl'),
                                    quantity=entry_quantity,
                                    product='NRML'
                                )
                                
                                # --- Multi-Broker SL Order ---
                                extra_sl_ids = {}
                                if self.extra_brokers and self.live_trading:
                                    logger.info(f"⚡ Placing COPY SL orders on {len(self.extra_brokers)} extra brokers...")
                                    
                                    with ThreadPoolExecutor(max_workers=3) as executor:
                                        futures = {
                                            executor.submit(
                                                self._place_extra_broker_sl_order, 
                                                name, service, 'PE', pe_strike, pe_signal.get('sl')
                                            ): name for name, service in self.extra_brokers.items()
                                        }
                                        
                                        for future in futures:
                                            try:
                                                result = future.result(timeout=5)
                                                if result:
                                                    b_name, b_order_id = result
                                                    extra_sl_ids[b_name] = b_order_id
                                            except Exception as e:
                                                logger.error(f"SL placement failed for {futures[future]}: {e}")
                                # -----------------------------

                                if sl_result['success']:
                                    sl_order_id = sl_result['order_id']
                                    logger.info(f"✅ PE SL order placed: {option_symbol} @ {pe_signal.get('sl'):.2f} | SL Order ID: {sl_order_id}")
                                    # Log SL order to Trade sheet
                                    excel_logger.log_trade(
                                        order_type='SL_ORDER',
                                        option_type='PE',
                                        strike=pe_strike,
                                        entry_price=pe_signal.get('sl'),
                                        target=pe_signal.get('target'),
                                        stop_loss=pe_signal.get('sl'),
                                        status='PLACED',
                                        order_id=sl_order_id,
                                        notes=f'SL Order Placed | Trigger: {pe_signal.get("sl"):.2f}'
                                    )
                                else:
                                    logger.error(f"❌ Failed to place PE SL order: {sl_result.get('error')}")
                                    # Log failure to Trade sheet
                                    excel_logger.log_trade(
                                        order_type='SL_ORDER',
                                        option_type='PE',
                                        strike=pe_strike,
                                        entry_price=pe_signal.get('sl'),
                                        target=pe_signal.get('target'),
                                        stop_loss=pe_signal.get('sl'),
                                        status='FAILED',
                                        notes=f'SL Order Failed: {sl_result.get("error")}'
                                    )
                            else:
                                logger.warning(f"Could not find option symbol for PE {pe_strike} - SL order not placed. Will monitor internally.")
                        except Exception as e:
                            logger.error(f"Error placing PE SL order: {e}", exc_info=True)
                    
                    self.active_trades['PE'] = {
                        'entry_price': pe_signal.get('entry_price'),
                        'entry_high': pe_signal.get('entry_high'),  # Reference high used for entry
                        'sl': pe_signal.get('sl'),
                        'target': pe_signal.get('target'),
                        'entry_time': datetime.now().isoformat(),
                        'order_id': order_id,
                        'sl_order_id': sl_order_id,  # Track SL order ID for modifications
                        'extra_sl_ids': extra_sl_ids, # Track SL IDs for other brokers
                        'token': strike_data.get('pe_token') if strike_data else None,  # type: ignore
                        'strike': pe_strike if strike_data else None,
                        'status': 'OPEN',
                        # Trailing SL state
                        'target_hit': False,  # Track if target was hit
                        'trailed_sl': pe_signal.get('sl'),  # Current trailed SL (starts at initial SL)
                        'sl_distance': pe_signal.get('entry_price') - pe_signal.get('sl')  # Distance between entry and SL
                    }
                    # Entry is already marked above to prevent race conditions
                    logger.info(f"🟢 PE {pe_strike} trade opened at {pe_signal.get('entry_price')} (Entry High: {pe_signal.get('entry_high')}, SL: {pe_signal.get('sl')}, Target: {pe_signal.get('target')}) | Order ID: {order_id if order_id else 'N/A'} | SL Order ID: {sl_order_id if sl_order_id else 'N/A'}")
                else:
                    logger.info(f"⏭️  PE signal detected but trade already OPEN - skipping to allow sequential entries (close existing trade first)")
    
    def log_signal(self, signals: Dict[str, Any]) -> None:
        """
        Log signal information for tracking.
        
        Args:
            signals: Entry signals
        """
        signal_entry = {
            'timestamp': signals.get('timestamp'),
            'ce_signal': signals.get('ce_signal'),
            'pe_signal': signals.get('pe_signal')
        }
        self.today_signals.append(signal_entry)
        
        # Keep only last 100 signals in memory
        if len(self.today_signals) > 100:
            self.today_signals.pop(0)
    
    def check_entry_signal_live(self, ce_token: int, pe_token: int, 
                               ce_high: float, pe_high: float) -> Dict[str, Any]:
        """
        Check for live entry signals using the strategy's check_entry_signal method.
        
        Reuses entry detection and SL/Target calculation logic from Intraday920Strategy.
        
        Args:
            ce_token: CE option token
            pe_token: PE option token
            ce_high: Reference high for CE entry (CE first 5-min high)
            pe_high: Reference high for PE entry (PE first 5-min high)
            
        Returns:
            Dictionary with entry signals from strategy.check_entry_signal()
        """
        try:
            # Use strategy's check_entry_signal method which:
            # 1. Fetches latest 5-min candles for both CE and PE
            # 2. Analyzes entry conditions for each side
            # 3. Calculates SL and Target using calculate_sl_for_entry
            # 4. Returns formatted signal data
            
            signals = self.strategy.check_entry_signal(
                ce_token=ce_token,
                pe_token=pe_token,
                ce_high=ce_high,
                pe_high=pe_high,
                symbol=self.symbol
            )
            
            return signals
            
        except Exception as e:
            logger.error(f"Error checking entry signals: {str(e)}")
            return {'success': False, 'error': str(e), 'timestamp': datetime.now().isoformat()}
    
    def place_buy_order(self, side: str, token: int, strike: int, entry_price: float, 
                        transaction_type: str = 'BUY') -> Optional[str]:
        """
        Place a buy order for CE or PE option when entry signal detected.
        
        Generic method that delegates to KiteService.place_option_order():
        1. Looks up the option trading symbol
        2. Fetches current market price
        3. Places market order via broker service
        
        Logs order placement to Excel Trade sheet.
        
        Args:
            side: 'CE' or 'PE'
            token: Option token (unused but kept for compatibility)
            strike: Strike price
            entry_price: Entry price from signal
            transaction_type: 'BUY' or 'SELL' (default: 'BUY')
            
        Returns:
            Order ID or None if failed
        """
        logger.info(f"place_buy_order called: {side} {strike} @ {entry_price:.2f} (live_trading={self.live_trading})")
        
        if not self.live_trading:
            demo_msg = f"DEMO: {transaction_type} {side} {strike} @ {entry_price:.2f}"
            logger.info(demo_msg)
            
            # Log DEMO order placement
            log_order_placement({
                'symbol': self.symbol,
                'side': f'{transaction_type} {side}',
                'strike': strike,
                'entry_price': f"{entry_price:.2f}",
                'order_type': 'MARKET',
                'mode': 'DEMO',
                'status': 'SUCCESS',
                'order_id': 'DEMO_ORDER',
                'sl': 'N/A',
                'target': 'N/A'
            })
            
            return "DEMO_ORDER"
        
        try:
            # Map transaction_type string to Kite constant
            transaction_type_map = {
                'BUY': self.kite.TRANSACTION_TYPE_BUY,
                'SELL': self.kite.TRANSACTION_TYPE_SELL
            }
            transaction_type_const = transaction_type_map.get(transaction_type, self.kite.TRANSACTION_TYPE_BUY)
            
            result = self.kite_service.place_option_order(
                symbol=self.symbol,
                strike=strike,
                option_type=side,
                transaction_type=transaction_type_const
            )
            
            # --- Multi-Broker Order Placement ---
            if self.extra_brokers and self.live_trading:
                logger.info(f"⚡ Placing COPY orders on {len(self.extra_brokers)} extra brokers...")
                for broker_name, service in self.extra_brokers.items():
                    try:
                        threading.Thread(
                            target=self._place_extra_broker_order,
                            args=(broker_name, service, side, strike, entry_price, transaction_type),
                            name=f"Order-{broker_name}-{side}-{strike}",
                            daemon=True
                        ).start()
                    except Exception as e:
                        logger.error(f"Failed to trigger {broker_name} order: {e}")
            # ------------------------------------
            
            if result['success']:
                logger.info(f"✅ {transaction_type} Order placed successfully. Order ID: {result['order_id']} | {side} {strike} @ {entry_price:.2f}")
                
                # Log successful live order placement
                log_order_placement({
                    'symbol': self.symbol,
                    'side': f'{transaction_type} {side}',
                    'strike': strike,
                    'entry_price': f"{entry_price:.2f}",
                    'order_type': 'MARKET',
                    'mode': 'LIVE',
                    'status': 'SUCCESS',
                    'order_id': result['order_id'],
                    'sl': 'N/A',
                    'target': 'N/A',
                    'details': f"Order placed successfully via KiteService"
                })
                
                return result['order_id']
            else:
                logger.error(f"❌ {transaction_type} Order failed: {result['error']}")
                
                # Log failed live order placement
                log_order_placement({
                    'symbol': self.symbol,
                    'side': f'{transaction_type} {side}',
                    'strike': strike,
                    'entry_price': f"{entry_price:.2f}",
                    'order_type': 'MARKET',
                    'mode': 'LIVE',
                    'status': 'FAILED',
                    'order_id': 'N/A',
                    'sl': 'N/A',
                    'target': 'N/A',
                    'error': result['error']
                })
                
                return None
                
        except Exception as e:
            logger.error(f"Error placing BUY order for {side} {strike}: {e}", exc_info=True)
            
            # Log exception during order placement
            log_order_placement({
                'symbol': self.symbol,
                'side': f'BUY {side}',
                'strike': strike,
                'entry_price': f"{entry_price:.2f}",
                'order_type': 'MARKET',
                'mode': 'LIVE',
                'status': 'EXCEPTION',
                'order_id': 'N/A',
                'sl': 'N/A',
                'target': 'N/A',
                'error': str(e)
            })
            
            return None
    
    def get_current_prices(self, tokens: List[int]) -> Dict[int, Optional[float]]:
        """
        Fetch multiple current prices (LTP) in single API call for efficiency.
        
        Generic method that delegates to KiteService for broker-agnostic price fetching.
        Reduces API overhead when monitoring multiple tokens.
        
        Args:
            tokens: List of instrument tokens to fetch
            
        Returns:
            Dictionary mapping token -> price
        """
        return self.kite_service.get_current_prices_batch(tokens)
    
    def get_current_price(self, token: int) -> Optional[float]:
        """
        Fetch current LTP (Last Traded Price) for a given token.
        
        Generic method that delegates to KiteService for broker-agnostic price fetching.
        
        Args:
            token: Instrument token
            
        Returns:
            Current price or None if failed
        """
        return self.kite_service.get_current_price(token)
    
    def _place_extra_broker_order(self, broker_name: str, service: Any, side: str, 
                                 strike: int, price: float, transaction_type: str) -> None:
        """
        Place generic buy/sell order on extra brokers.
        Executes in a separate thread.
        """
        try:
            qty = self.kite_service.get_lot_size(self.symbol)
            txn_type = 'BUY' if transaction_type == 'BUY' else 'SELL'
            
            logger.info(f"⚡ [{broker_name}] Attempting {txn_type} {side} {strike} x {qty}...")

            response = None
            
            if broker_name == 'KOTAK':
                # Kotak uses 'B' for BUY and 'S' for SELL
                k_txn = 'B' if txn_type == 'BUY' else 'S'
                
                # Kotak service handles symbol construction internally
                response = service.place_option_order(
                    symbol=self.symbol,
                    strike=strike,
                    option_type=side,
                    transaction_type=k_txn,
                    quantity=qty
                )
            
            elif broker_name == 'DHAN':
                # Dhan needs numeric security ID for options
                sec_id = service.get_option_security_id(self.symbol, strike, side)
                if sec_id:
                    response = service.place_order(
                        security_id=sec_id,
                        transaction_type=txn_type,
                        quantity=qty,
                        order_type='MARKET',
                        product_type='INTRADAY',
                        exchange_segment='NSE_FNO'
                    )
                else:
                    logger.error(f"❌ [{broker_name}] Could not resolve security ID for {side} {strike}")
                    return

            elif broker_name == 'FYERS':
                # Fyers needs NSE:SYMBOL format e.g. NSE:NIFTY24JAN21500CE
                kite_symbol = self._get_option_symbol(self.symbol, strike, side)
                if kite_symbol:
                    fyers_symbol = f"NSE:{kite_symbol}"
                    # Fyers side: 1 (Buy), -1 (Sell)
                    f_side = 1 if txn_type == 'BUY' else -1
                    response = service.place_order(
                        symbol=fyers_symbol,
                        side=f_side,
                        quantity=qty,
                        order_type=2, # MARKET
                        product_type='INTRADAY'
                    )
                else:
                    logger.error(f"❌ [{broker_name}] Could not resolve symbol for {side} {strike}")
                    return

            # Log result
            if response and isinstance(response, dict):
                if response.get('success') or response.get('stat') == 'Ok' or 'nOrdNo' in response:
                    logger.info(f"✅ [{broker_name}] Order Placed Successfully: {response}")
                else:
                    logger.error(f"❌ [{broker_name}] Order Failed: {response}")
            else:
                logger.error(f"❌ [{broker_name}] Invalid or No Response: {response}")

        except Exception as e:
            logger.error(f"❌ [{broker_name}] Exception during order placement: {e}", exc_info=True)

    def _place_extra_broker_sl_order(self, broker_name: str, service: Any, side: str, 
                                    strike: int, trigger_price: float) -> Optional[tuple]:
        """
        Place generic SL (Sell) order on extra brokers.
        Returns: (broker_name, order_id) or None
        """
        try:
            qty = self.kite_service.get_lot_size(self.symbol)
            logger.info(f"⚡ [{broker_name}] Attempting SL SELL {side} {strike} @ {trigger_price}...")
            
            response = None
            order_id = None

            if broker_name == 'KOTAK':
                # Kotak now stores SL-M orders via place_stoploss_order (which wraps place_order)
                # Ensure implementation exists in service
                # If using generic place_option_stoploss_order which I should add to service wrapper or just use here
                
                # Kotak generic place_option_order doesn't support trigger yet in wrapper unless I update it
                # I updated KotakOrderService.place_option_stoploss_order.
                
                # Using the new helper I added (or assumed to add) or just place_order
                # Let's use the object's method if available, else standard fallback
                
                if hasattr(service, 'place_option_stoploss_order'):
                     response = service.place_option_stoploss_order(
                        symbol=self.symbol,
                        strike=strike,
                        option_type=side,
                        trigger_price=trigger_price,
                        quantity=qty,
                        transaction_type='S'
                    )
                else:
                    # Fallback if service not reloaded yet in memory (should be though)
                    logger.warning(f"[{broker_name}] Service missing place_option_stoploss_order")
                    return None
            
            elif broker_name == 'DHAN':
                # Dhan needs security ID
                sec_id = service.get_option_security_id(self.symbol, strike, side)
                if sec_id:
                    response = service.place_stoploss_order(
                        security_id=sec_id,
                        trigger_price=trigger_price,
                        quantity=qty,
                        product_type='INTRADAY',
                        exchange_segment='NSE_FNO'
                    )
                else:
                    logger.error(f"❌ [{broker_name}] Could not resolve security ID for SL {side} {strike}")
                    return None

            elif broker_name == 'FYERS':
                # Fyers needs NSE:SYMBOL
                kite_symbol = self._get_option_symbol(self.symbol, strike, side)
                if kite_symbol:
                    fyers_symbol = f"NSE:{kite_symbol}"
                    response = service.place_stoploss_order(
                        symbol=fyers_symbol,
                        trigger_price=trigger_price,
                        quantity=qty,
                        product_type='INTRADAY'
                    )
                else:
                    logger.error(f"❌ [{broker_name}] Could not resolve symbol for SL {side} {strike}")
                    return None

            # Log result
            if response and isinstance(response, dict):
                if response.get('success') or response.get('stat') == 'Ok' or 'nOrdNo' in response:
                    logger.info(f"✅ [{broker_name}] SL Order Placed Successfully: {response}")
                    order_id = response.get('order_id') or response.get('nOrdNo') or response.get('id')
                    return (broker_name, str(order_id))
                else:
                    logger.error(f"❌ [{broker_name}] SL Order Failed: {response}")
            else:
                logger.error(f"❌ [{broker_name}] SL Order Invalid/No Response: {response}")

        except Exception as e:
            logger.error(f"❌ [{broker_name}] SL Order placement failed: {e}", exc_info=True)
        
        return None

    def _modify_extra_broker_sl_order(self, broker_name: str, service: Any, order_id: str, 
                                     new_trigger_price: float) -> bool:
        """
        Modify SL order on extra broker.
        """
        try:
            logger.info(f"⚡ [{broker_name}] Modifying SL Order {order_id} -> {new_trigger_price}...")
            
            response = None
            qty = self.kite_service.get_lot_size(self.symbol) # Sometimes needed
            
            if broker_name == 'KOTAK':
                # Kotak modify_order(order_id, price, quantity, trigger_price)
                if hasattr(service, 'modify_order'):
                    response = service.modify_order(
                        order_id=order_id,
                        trigger_price=new_trigger_price
                    )
            
            elif broker_name == 'DHAN':
                # Dhan modify_order(order_id, quantity, price, order_type, trigger_price, validity)
                if hasattr(service, 'modify_order'):
                    response = service.modify_order(
                        order_id=order_id,
                        order_type='STOP_LOSS_MARKET',
                        trigger_price=new_trigger_price,
                        quantity=qty,
                        validity='DAY'
                    )

            elif broker_name == 'FYERS':
                # Fyers modify_order doesn't support trigger_price changes
                # Use cancel + replace strategy instead
                try:
                    cancel_resp = service.cancel_order(order_id=order_id)
                    if cancel_resp.get('success'):
                        logger.info(f"[FYERS] Old SL cancelled, placing new SL @ {new_trigger_price}")
                        # Get the trade to find the symbol
                        for side_key, trade_data in self.active_trades.items():
                            if trade_data.get('extra_sl_ids', {}).get('FYERS') == order_id:
                                strike = trade_data.get('strike')
                                kite_symbol = self._get_option_symbol(self.symbol, strike, side_key)
                                if kite_symbol:
                                    fyers_symbol = f"NSE:{kite_symbol}"
                                    new_resp = service.place_stoploss_order(
                                        symbol=fyers_symbol,
                                        trigger_price=new_trigger_price,
                                        quantity=qty,
                                        product_type='INTRADAY'
                                    )
                                    if new_resp and new_resp.get('success'):
                                        new_oid = new_resp.get('order_id') or new_resp.get('id')
                                        trade_data['extra_sl_ids']['FYERS'] = str(new_oid)
                                        response = new_resp
                                    else:
                                        response = new_resp
                                break
                    else:
                        logger.error(f"[FYERS] Failed to cancel old SL: {cancel_resp}")
                        response = cancel_resp
                except Exception as fyers_e:
                    logger.error(f"[FYERS] Cancel+Replace SL failed: {fyers_e}")
                    response = {'success': False, 'error': str(fyers_e)}
            
            # Check success
            if response and isinstance(response, dict):
                 if response.get('success') or response.get('stat') == 'Ok':
                     logger.info(f"✅ [{broker_name}] SL Modified: {new_trigger_price}")
                     return True
                 else:
                     logger.error(f"❌ [{broker_name}] Modification Failed: {response}")
            
        except Exception as e:
            logger.error(f"❌ [{broker_name}] Modify Exception: {e}")
            
        return False

    def _cancel_pending_sl_orders(self, trade: Dict[str, Any], side: str) -> None:
        """
        Cancel all pending SL orders (Kite + extra brokers) for a trade being closed.
        Must be called when exiting a trade via SL hit, trailing SL hit, or any exit path.
        Prevents stale SL orders from triggering on subsequent positions.
        
        Args:
            trade: Trade dictionary containing sl_order_id and extra_sl_ids
            side: 'CE' or 'PE' (for logging)
        """
        # 1. Cancel Kite SL order
        sl_order_id = trade.get('sl_order_id')
        if sl_order_id and self.live_trading:
            try:
                cancel_result = self.kite_service.cancel_order(order_id=sl_order_id)
                if cancel_result.get('success', False):
                    logger.info(f"✅ {side} Kite SL order cancelled: {sl_order_id}")
                else:
                    logger.warning(f"⚠️ {side} Kite SL order cancel response: {cancel_result}")
            except Exception as e:
                # Order may already be completed/cancelled - this is expected
                logger.debug(f"{side} Kite SL cancel (may already be done): {e}")
        
        # 2. Cancel extra broker SL orders
        extra_sl_ids = trade.get('extra_sl_ids', {})
        if extra_sl_ids and self.live_trading:
            for b_name, b_sl_id in extra_sl_ids.items():
                if b_name in self.extra_brokers:
                    try:
                        service_obj = self.extra_brokers[b_name]
                        if hasattr(service_obj, 'cancel_order'):
                            threading.Thread(
                                target=service_obj.cancel_order,
                                args=(b_sl_id,),
                                name=f"CancelSL-{b_name}-{side}",
                                daemon=True
                            ).start()
                            logger.info(f"✅ [{b_name}] SL cancel triggered for {side}: {b_sl_id}")
                        else:
                            logger.warning(f"[{b_name}] No cancel_order method found")
                    except Exception as e:
                        logger.error(f"Failed to cancel {b_name} SL {b_sl_id}: {e}")

    def _get_option_symbol(self, symbol: str, strike: int, option_type: str) -> Optional[str]:
        """
        Get the trading symbol for an option contract.
        e.g., 'NIFTY25D26C25000' for NIFTY, strike 25000, CE, expiry 26-DEC-2025
        
        Filters by nearest expiry >= today to avoid returning wrong contract.
        
        Args:
            symbol: Underlying symbol (NIFTY, BANKNIFTY, etc.)
            strike: Strike price
            option_type: 'CE' or 'PE'
            
        Returns:
            Trading symbol or None if not found
        """
        try:
            logger.debug(f"Looking up option symbol for {symbol} {strike} {option_type}")
            
            # Use kite_service to look up the option symbol
            if hasattr(self.kite_service, '_nfo_instruments_cache'):
                nfo_instruments = self.kite_service._nfo_instruments_cache
                if nfo_instruments:
                    logger.debug(f"Searching in {len(nfo_instruments)} NFO instruments")
                    
                    from datetime import date as date_type
                    today = date_type.today()
                    
                    # Collect all matching instruments and pick nearest expiry
                    candidates = []
                    for instrument in nfo_instruments:
                        if (instrument.get('name') == symbol and 
                            instrument.get('strike') == strike and
                            instrument.get('instrument_type') == option_type):
                            expiry = instrument.get('expiry')
                            if expiry:
                                if hasattr(expiry, 'date'):
                                    expiry = expiry.date()
                                if expiry >= today:
                                    candidates.append((expiry, instrument.get('tradingsymbol')))
                    
                    if candidates:
                        # Sort by expiry and return the nearest one
                        candidates.sort(key=lambda x: x[0])
                        nearest_expiry, trading_symbol = candidates[0]
                        logger.info(f"✅ Found option symbol: {trading_symbol} (expiry: {nearest_expiry}) for {symbol} {strike} {option_type}")
                        return trading_symbol
                    
                    logger.warning(f"No match found in NFO cache for {symbol} {strike} {option_type}")
                else:
                    logger.warning("NFO instruments cache is empty")
            else:
                logger.warning("KiteService does not have _nfo_instruments_cache")
            
            # Fallback
            logger.warning(f"Could not find option symbol for {symbol} {strike} {option_type} - SL order will not be placed")
            return None
            
        except Exception as e:
            logger.error(f"Error getting option symbol for {symbol} {strike} {option_type}: {e}", exc_info=True)
            return None
    
    def check_sl_target_for_active_trades(self, check_timestamp: Optional[datetime] = None) -> None:
        """
        Monitor active trades and check if SL or Target has been hit.
        Implements 1:2 with Trailing SL logic:
        
        1. Before target hit: Exit if price <= initial SL
        2. When target hit: Move SL to entry price (lock in breakeven)
        3. After target hit: Trail SL by sl_distance for every sl_distance price moves above entry
        4. Exit when trailed SL is hit
        
        Automatically places SELL orders when conditions are met.
        Uses batch price fetching for efficiency when monitoring multiple trades.
        
        Args:
            check_timestamp: Timestamp when this check was initiated (for accurate logging/Excel tracking)
        """
        if check_timestamp is None:
            check_timestamp = datetime.now()
        if not self.active_trades:
            return
        
        # Fetch all prices at once (efficient batch operation)
        tokens = [t.get('token') for t in self.active_trades.values() if t.get('token')]
        prices = self.get_current_prices(tokens) if tokens else {}
        
        trades_to_close = []
        
        with self.active_trades_lock:
            for side, trade in list(self.active_trades.items()):
                if trade.get('status') != 'OPEN':
                    continue
                
                token = trade.get('token')
                strike = trade.get('strike')
                
                if not token or not strike or token not in prices:
                    if token and token not in prices:
                        logger.warning(f"Failed to fetch price for {side} {strike}")
                    continue
                
                # Get current price from batch fetch results
                current_price = prices[token]
                
                if current_price is None:
                    logger.warning(f"No price available for {side} {strike}")
                    continue
                
                entry_price = trade.get('entry_price', 0)
                initial_sl = trade.get('sl', 0)
                target = trade.get('target', 0)
                target_hit = trade.get('target_hit', False)
                trailed_sl = trade.get('trailed_sl', initial_sl)
                sl_distance = trade.get('sl_distance', entry_price - initial_sl)
                
                # === 1:2 WITH TRAILING SL LOGIC ===
                if target_hit:
                    # Target already hit - implement trailing SL
                    
                    # === NEW 1:2 TRAILING LOGIC (10 POINT STEP) ===
                    # Formula: New_SL = Entry + int((Price - Target) / 10) * 10
                    # At Target (Price=Target): Diff=0, SL = Entry + 0 = Entry
                    # At Target+10 (Price=Target+10): Diff=10, SL = Entry + 10
                    
                    price_above_target = current_price - target
                    TRAIL_STEP = 10.0
                    
                    # Only trail if price is at or above target (which it should be if target_hit is True)
                    if price_above_target >= 0:
                        # Calculate how many full 10-point steps we are above the target
                        steps_above_target = int(price_above_target / TRAIL_STEP)
                        
                        # New SL is Entry + (Steps * 10)
                        new_trailed_sl = entry_price + (steps_above_target * TRAIL_STEP)
                        
                        if new_trailed_sl > trailed_sl:
                            old_sl = trailed_sl
                            trailed_sl = new_trailed_sl
                            trade['trailed_sl'] = trailed_sl
                            logger.info(f"📈 {side} Trailing SL updated: {trailed_sl:.2f} (price: {current_price:.2f}, target: {target:.2f}, steps: {steps_above_target})")
                            
                            # Modify SL order on broker if live trading
                            sl_order_id = trade.get('sl_order_id')
                            if self.live_trading:
                                # 1. Modify Kite SL
                                if sl_order_id:
                                    try:
                                        modify_result = self.kite_service.modify_stoploss_order(
                                            order_id=sl_order_id,
                                            new_trigger_price=trailed_sl
                                        )
                                        if modify_result['success']:
                                            logger.info(f"✅ {side} SL order modified on broker: {sl_order_id} -> Trigger: {trailed_sl:.2f}")
                                            # Log SL update to Trade sheet
                                            excel_logger.log_trade(
                                                order_type='SL_UPDATE',
                                                option_type=side,
                                                strike=strike,
                                                entry_price=entry_price,
                                                current_price=current_price,
                                                target=target,
                                                stop_loss=trailed_sl,
                                                pnl=current_price - entry_price,
                                                status='SL_TRAILED',
                                                order_id=sl_order_id,
                                                notes=f'SL Trailed from {old_sl:.2f} to {trailed_sl:.2f}'
                                            )
                                        else:
                                            logger.error(f"❌ Failed to modify {side} SL order: {modify_result.get('error')}")
                                    except Exception as e:
                                        logger.error(f"Error modifying {side} SL order: {e}", exc_info=True)
                                
                                # 2. Modify Extra Broker SLs
                                if trade.get('extra_sl_ids'):
                                    for b_name, b_order_id in trade['extra_sl_ids'].items():
                                        if b_name in self.extra_brokers:
                                            try:
                                                threading.Thread(
                                                    target=self._modify_extra_broker_sl_order,
                                                    args=(b_name, self.extra_brokers[b_name], b_order_id, trailed_sl),
                                                    name=f"ModSL-{b_name}-{side}",
                                                    daemon=True
                                                ).start()
                                            except Exception as e:
                                                logger.error(f"Failed to trigger {b_name} SL modification: {e}")
                            
                            # Log to Signal Checks sheet
                            excel_logger.log_sl_target_check(
                                timestamp=check_timestamp,
                                side=side,
                                strike=strike,
                                current_price=current_price,
                                entry_price=entry_price,
                                initial_sl=initial_sl,
                                target=target,
                                target_hit=True,
                                trailed_sl=trailed_sl,
                                check_reason="TRAILING"
                            )
                            
                            # Log to Excel
                            profit = current_price - entry_price
                            excel_logger.log_trailing_sl_update(
                                option_type=side,
                                strike=strike,
                                old_sl=old_sl,
                                new_sl=trailed_sl,
                                current_price=current_price,
                                profit=profit
                            )
                    else:
                        # Price pulled back below entry - ensure SL stays at entry (lock in gain)
                        if trailed_sl < entry_price:
                            trailed_sl = entry_price
                            trade['trailed_sl'] = trailed_sl
                            logger.info(f"🔒 {side} SL locked at entry: {entry_price:.2f} (price pulled back to {current_price:.2f})")
                    
                    # Check if trailed SL is hit
                    if current_price <= trailed_sl:
                        logger.info(f"🔴 TRAILED SL HIT for {side}: Current {current_price:.2f} <= Trailed SL {trailed_sl:.2f}")
                        
                        # Log to Signal Checks sheet
                        excel_logger.log_sl_target_check(
                            timestamp=check_timestamp,
                            side=side,
                            strike=strike,
                            current_price=current_price,
                            entry_price=entry_price,
                            initial_sl=initial_sl,
                            target=target,
                            target_hit=True,
                            trailed_sl=trailed_sl,
                            check_reason="SL_HIT"
                        )
                        
                        # Log to Excel
                        pnl = current_price - entry_price
                        excel_logger.log_trade(
                            order_type='SELL',
                            option_type=side,
                            strike=strike,
                            entry_price=entry_price,
                            current_price=current_price,
                            target=target,
                            stop_loss=trailed_sl,
                            pnl=pnl,
                            status='TRAILED_SL_HIT',
                            notes=f"Trailed SL Hit at {trailed_sl:.2f} | Entry: {entry_price:.2f} | P&L: {pnl:+.2f}"
                        )
                        
                        # Cancel pending SL orders before placing market sell
                        self._cancel_pending_sl_orders(trade, side)
                        
                        # Place sell order
                        order_id = self.place_sell_order(
                            side=side,
                            strike=strike,
                            exit_price=current_price,
                            exit_reason=f"Trailed SL Hit ({trailed_sl:.2f})"
                        )
                        
                        # Close trade
                        self.close_trade(side, current_price, f"Trailed SL Hit ({trailed_sl:.2f})")
                        trades_to_close.append(side)
                    else:
                        # Still in trade - log status
                        pnl = current_price - entry_price
                        pnl_pct = (pnl / entry_price * 100) if entry_price else 0
                        logger.debug(f"📊 {side} Trade (Target Hit): Entry {entry_price:.2f}, Current {current_price:.2f}, Trailed SL {trailed_sl:.2f} | PnL: {pnl:.2f} ({pnl_pct:.2f}%)")
                        
                else:
                    # Target not yet hit - check for initial SL or target hit
                    
                    # Check if SL hit (before target)
                    if current_price <= initial_sl:
                        logger.info(f"🔴 SL HIT for {side}: Current {current_price:.2f} <= SL {initial_sl:.2f}")
                        
                        # Log to Signal Checks sheet
                        excel_logger.log_sl_target_check(
                            timestamp=check_timestamp,
                            side=side,
                            strike=strike,
                            current_price=current_price,
                            entry_price=entry_price,
                            initial_sl=initial_sl,
                            target=target,
                            target_hit=False,
                            check_reason="SL_HIT"
                        )
                        
                        # Log to Excel
                        pnl = current_price - entry_price
                        excel_logger.log_trade(
                            order_type='SELL',
                            option_type=side,
                            strike=strike,
                            entry_price=entry_price,
                            current_price=current_price,
                            target=target,
                            stop_loss=initial_sl,
                            pnl=pnl,
                            status='SL_HIT',
                            notes=f"Initial SL Hit at {initial_sl:.2f} | Entry: {entry_price:.2f} | P&L: {pnl:+.2f}"
                        )
                        
                        # Cancel pending SL orders before placing market sell
                        self._cancel_pending_sl_orders(trade, side)
                        
                        # Place sell order
                        order_id = self.place_sell_order(
                            side=side,
                            strike=strike,
                            exit_price=current_price,
                            exit_reason="SL Hit"
                        )
                        
                        # Close trade
                        self.close_trade(side, current_price, "SL Hit")
                        trades_to_close.append(side)
                        
                    # Check if Target hit (activate trailing SL)
                    elif current_price >= target:
                        logger.info(f"🎯 TARGET HIT for {side}: Current {current_price:.2f} >= Target {target:.2f}")
                        logger.info(f"🔄 Trailing SL activated for {side} - SL moved to entry {entry_price:.2f}")
                        
                        # Log to Signal Checks sheet
                        excel_logger.log_sl_target_check(
                            timestamp=check_timestamp,
                            side=side,
                            strike=strike,
                            current_price=current_price,
                            entry_price=entry_price,
                            initial_sl=initial_sl,
                            target=target,
                            target_hit=False,
                            check_reason="TARGET_HIT"
                        )
                        
                        # Log to Excel
                        pnl = current_price - entry_price
                        excel_logger.log_trade(
                            order_type='UPDATE',
                            option_type=side,
                            strike=strike,
                            entry_price=entry_price,
                            current_price=current_price,
                            target=target,
                            stop_loss=entry_price,  # SL moved to entry
                            pnl=pnl,
                            status='TARGET_HIT',
                            notes=f"Target Hit at {target:.2f} | Trailing SL activated | SL moved to entry {entry_price:.2f} | P&L: {pnl:+.2f}"
                        )
                        
                        # Activate trailing SL - move SL to entry price
                        trade['target_hit'] = True
                        trade['trailed_sl'] = entry_price

                        # === FIX: Update Broker SL to Entry Price ===
                        sl_order_id = trade.get('sl_order_id')
                        if self.live_trading:
                            # 1. Update Kite SL
                            if sl_order_id:
                                try:
                                    modify_result = self.kite_service.modify_stoploss_order(
                                        order_id=sl_order_id,
                                        new_trigger_price=entry_price
                                    )
                                    if modify_result['success']:
                                        logger.info(f"✅ {side} Broker SL moved to Entry: {sl_order_id} -> Trigger: {entry_price:.2f}")
                                        
                                        # Log SL update to Trade sheet (Breakeven move)
                                        excel_logger.log_trade(
                                            order_type='SL_UPDATE',
                                            option_type=side,
                                            strike=strike,
                                            entry_price=entry_price,
                                            current_price=current_price,
                                            target=target,
                                            stop_loss=entry_price,
                                            pnl=pnl,
                                            status='SL_MOVED_TO_ENTRY',
                                            order_id=sl_order_id,
                                            notes=f'Target Hit: SL moved to Breakeven {entry_price:.2f}'
                                        )
                                    else:
                                        logger.error(f"❌ Failed to move {side} SL to Entry: {modify_result.get('error')}")
                                except Exception as e:
                                    logger.error(f"Error moving {side} SL to Entry: {e}", exc_info=True)
                            
                            # 2. Update Extra Broker SLs
                            if trade.get('extra_sl_ids'):
                                for b_name, b_order_id in trade['extra_sl_ids'].items():
                                    if b_name in self.extra_brokers:
                                        try:
                                            threading.Thread(
                                                target=self._modify_extra_broker_sl_order,
                                                args=(b_name, self.extra_brokers[b_name], b_order_id, entry_price),
                                                name=f"ModSL-{b_name}-{side}",
                                                daemon=True
                                            ).start()
                                        except Exception as e:
                                            logger.error(f"Failed to trigger {b_name} SL modification: {e}")
                        # ============================================
                        
                        # Don't exit yet - continue with trailing SL
                        
                    else:
                        # Still waiting for target or SL - log status
                        pnl = current_price - entry_price
                        pnl_pct = (pnl / entry_price * 100) if entry_price else 0
                        logger.debug(f"📊 {side} Trade Status: Entry {entry_price:.2f}, Current {current_price:.2f}, SL {initial_sl:.2f}, Target {target:.2f} | PnL: {pnl:.2f} ({pnl_pct:.2f}%)")
        
        # Mark closed trades as CLOSED (don't delete - allows next entry to detect status)
        with self.active_trades_lock:
            for side in trades_to_close:
                if side in self.active_trades:
                    self.active_trades[side]['status'] = 'CLOSED'
                    logger.info(f"📋 {side} trade marked as CLOSED - allowing new entry on next signal")
    
    def place_sell_order(self, side: str, strike: int, exit_price: float, exit_reason: str = "Manual Exit",
                         transaction_type: str = 'SELL') -> Optional[str]:
        """
        Place a sell order for CE or PE option to exit trade.
        
        Generic method that delegates to KiteService.place_option_order():
        1. Looks up the option trading symbol
        2. Fetches current market price
        3. Places market order via broker service
        
        Logs order placement to Excel Trade sheet.
        
        Args:
            side: 'CE' or 'PE'
            strike: Strike price
            exit_price: Exit price
            exit_reason: Reason for exit (Target Hit, SL Hit, Manual, etc.)
            transaction_type: 'BUY' or 'SELL' (default: 'SELL')
            
        Returns:
            Order ID or None if failed
        """
        logger.info(f"place_sell_order called: {side} {strike} @ {exit_price:.2f} | Reason: {exit_reason} (live_trading={self.live_trading})")
        
        if not self.live_trading:
            demo_msg = f"DEMO: {transaction_type} {side} {strike} @ {exit_price:.2f} | {exit_reason}"
            logger.info(demo_msg)
            
            # Log DEMO sell order placement
            log_order_placement({
                'symbol': self.symbol,
                'side': f'{transaction_type} {side}',
                'strike': strike,
                'entry_price': f"{exit_price:.2f}",
                'order_type': 'MARKET',
                'mode': 'DEMO',
                'status': 'SUCCESS',
                'order_id': 'DEMO_ORDER',
                'sl': 'N/A',
                'target': 'N/A',
                'details': f"Exit Reason: {exit_reason}"
            })
            
            return "DEMO_ORDER"
        
        try:
            # Map transaction_type string to Kite constant
            transaction_type_map = {
                'BUY': self.kite.TRANSACTION_TYPE_BUY,
                'SELL': self.kite.TRANSACTION_TYPE_SELL
            }
            transaction_type_const = transaction_type_map.get(transaction_type, self.kite.TRANSACTION_TYPE_SELL)
            
            result = self.kite_service.place_option_order(
                symbol=self.symbol,
                strike=strike,
                option_type=side,
                transaction_type=transaction_type_const
            )
            
            # --- Multi-Broker EXIT Order ---
            if self.extra_brokers and self.live_trading:
                logger.info(f"⚡ Placing COPY EXIT orders on {len(self.extra_brokers)} extra brokers...")
                
                # 1. Place Market Exit
                for broker_name, service in self.extra_brokers.items():
                    try:
                        threading.Thread(
                            target=self._place_extra_broker_order,
                            args=(broker_name, service, side, strike, exit_price, transaction_type),
                            name=f"Exit-{broker_name}-{side}-{strike}",
                            daemon=True
                        ).start()
                    except Exception as e:
                        logger.error(f"Failed to trigger {broker_name} exit order: {e}")
                
                # 2. Cancel Pending SL Orders
                try:
                    trade = self.active_trades.get(side)
                    if trade and trade.get('extra_sl_ids'):
                        logger.info(f"Cancelling pending SL orders for {side}...")
                        for b_name, b_sl_id in trade['extra_sl_ids'].items():
                            if b_name in self.extra_brokers:
                                try:
                                    service_obj = self.extra_brokers[b_name]
                                    # Call cancel generic or specific?
                                    # Services usually have cancel_order(order_id)
                                    if hasattr(service_obj, 'cancel_order'):
                                        threading.Thread(
                                            target=service_obj.cancel_order,
                                            args=(b_sl_id,),
                                            name=f"CancelSL-{b_name}",
                                            daemon=True
                                        ).start()
                                    else:
                                        logger.warning(f"[{b_name}] No cancel_order method found")
                                except Exception as e:
                                    logger.error(f"Failed to cancel {b_name} SL {b_sl_id}: {e}")
                except Exception as e:
                    logger.error(f"Error coordinating SL cancellation: {e}")
            # -------------------------------
            
            if result['success']:
                logger.info(f"✅ {transaction_type} Order placed successfully. Order ID: {result['order_id']} | {side} {strike} @ {exit_price:.2f} | {exit_reason}")
                
                # Log successful live sell order placement
                log_order_placement({
                    'symbol': self.symbol,
                    'side': f'{transaction_type} {side}',
                    'strike': strike,
                    'entry_price': f"{exit_price:.2f}",
                    'order_type': 'MARKET',
                    'mode': 'LIVE',
                    'status': 'SUCCESS',
                    'order_id': result['order_id'],
                    'sl': 'N/A',
                    'target': 'N/A',
                    'details': f"Exit Reason: {exit_reason}"
                })
                
                return result['order_id']
            else:
                logger.error(f"❌ SELL Order failed: {result['error']} ({exit_reason})")
                
                # Log failed live sell order placement
                log_order_placement({
                    'symbol': self.symbol,
                    'side': f'SELL {side}',
                    'strike': strike,
                    'entry_price': f"{exit_price:.2f}",
                    'order_type': 'MARKET',
                    'mode': 'LIVE',
                    'status': 'FAILED',
                    'order_id': 'N/A',
                    'sl': 'N/A',
                    'target': 'N/A',
                    'error': result['error'],
                    'details': f"Exit Reason: {exit_reason}"
                })
                
                return None
                
        except Exception as e:
            logger.error(f"Error placing SELL order for {side} {strike}: {e}", exc_info=True)
            
            # Log exception during sell order placement
            log_order_placement({
                'symbol': self.symbol,
                'side': f'SELL {side}',
                'strike': strike,
                'entry_price': f"{exit_price:.2f}",
                'order_type': 'MARKET',
                'mode': 'LIVE',
                'status': 'EXCEPTION',
                'order_id': 'N/A',
                'sl': 'N/A',
                'target': 'N/A',
                'error': str(e),
                'details': f"Exit Reason: {exit_reason}"
            })
            
            return None
    
    def _entry_monitor_loop(self) -> None:
        """
        Entry signal monitoring loop - runs in background thread.
        Checked only at 5-minute marks (9:15, 9:20, 9:25, ..., 3:15, 3:20)
        """
        logger.info(f"Starting ENTRY signal monitoring for {self.symbol} [Instance ID: {id(self)}]")

        
        while self.is_entry_monitoring:
            try:
                # Check if within market hours
                if not self.is_market_hours():
                    if self.is_market_day():
                        logger.debug("Outside market hours, waiting...")
                    time_module.sleep(1)
                    continue
                
                # ====== ENTRY SIGNAL CHECK (5-minute marks only) ======
                if self.should_check_entry_signal():
                    # Use normalized 5-minute interval time set by should_check_entry_signal
                    check_timestamp = self.last_entry_check_time or datetime.now()
                    logger.info(f"[5-min Check] Checking entry signals for {self.symbol} at {check_timestamp.strftime('%H:%M:%S')}")
                    
                    # Fetch live data (timeout-protected)
                    live_data = self._fetch_live_data_with_timeout()
                    
                    # ALWAYS log the check attempt to Signal Checks sheet, even if it fails
                    if not live_data.get('success'):
                        error_msg = live_data.get('error', 'Unknown error')
                        logger.warning(f"Entry signal check failed at {check_timestamp.strftime('%H:%M:%S')}: {error_msg}")
                        
                        # Log the failed check to Excel so we can see the gap
                        excel_logger.log_signal_check(
                            timestamp=check_timestamp,
                            ce_prev_high=None,
                            ce_prev_low=None,
                            pe_prev_high=None,
                            pe_prev_low=None,
                            ce_signal=False,
                            pe_signal=False,
                            notes=f"Data fetch failed: {error_msg}"
                        )
                    
                    if live_data.get('success'):
                        # Extract strike info
                        high_strike = live_data.get('high_strike', {})
                        low_strike = live_data.get('low_strike', {})
                        
                        logger.info(f"High Strike - CE Token: {high_strike.get('ce_token')}, PE Token: {high_strike.get('pe_token')}, CE High: {high_strike.get('ce_high')}, PE High: {high_strike.get('pe_high')}")
                        logger.info(f"Low Strike - CE Token: {low_strike.get('ce_token')}, PE Token: {low_strike.get('pe_token')}, CE High: {low_strike.get('ce_high')}, PE High: {low_strike.get('pe_high')}")
                        
                        # Initialize data for Excel logging
                        has_ce_signal = False
                        has_pe_signal = False
                        ce_entry_price = None
                        pe_entry_price = None
                        ce_sl = None
                        pe_sl = None
                        ce_target = None
                        pe_target = None
                        ce_high_val = high_strike.get('ce_high')
                        pe_high_val = high_strike.get('pe_high')
                        
                        if high_strike.get('success') and low_strike.get('success'):
                            # Check signals for high strike
                            high_signals = self.check_entry_signal_live(
                                high_strike.get('ce_token'),
                                high_strike.get('pe_token'),
                                high_strike.get('ce_high'),
                                high_strike.get('pe_high')
                            )
                            
                            if high_signals.get('success'):
                                ce_sig = high_signals.get('ce_signal', {})
                                pe_sig = high_signals.get('pe_signal', {})
                                
                                # Log what was received
                                logger.info(f"HIGH STRIKE Signals - CE has_signal: {ce_sig.get('has_signal')}, PE has_signal: {pe_sig.get('has_signal')}")
                                if not ce_sig.get('has_signal'):
                                    logger.info(f"CE No Signal Reason: {ce_sig.get('reason', 'Unknown')}")
                                if not pe_sig.get('has_signal'):
                                    logger.info(f"PE No Signal Reason: {pe_sig.get('reason', 'Unknown')}")
                                
                                # Update Excel logging data if signals exist
                                if ce_sig.get('has_signal'):
                                    has_ce_signal = True
                                    ce_entry_price = ce_sig.get('entry_price')
                                    ce_sl = ce_sig.get('sl')
                                    ce_target = ce_sig.get('target')
                                
                                if pe_sig.get('has_signal'):
                                    has_pe_signal = True
                                    pe_entry_price = pe_sig.get('entry_price')
                                    pe_sl = pe_sig.get('sl')
                                    pe_target = pe_sig.get('target')
                                
                                # Only update trades if there are actual signals
                                if ce_sig.get('has_signal') or pe_sig.get('has_signal'):
                                    self.update_active_trades(high_signals, strike_data=high_strike)
                                    self.log_signal(high_signals)
                                
                                if ce_sig.get('has_signal'):
                                    logger.info(f"📊 HIGH STRIKE CE SIGNAL: Entry {ce_sig.get('entry_price')}, SL {ce_sig.get('sl')}, Target {ce_sig.get('target')} | ✅ Order placed")
                                
                                if pe_sig.get('has_signal'):
                                    logger.info(f"📊 HIGH STRIKE PE SIGNAL: Entry {pe_sig.get('entry_price')}, SL {pe_sig.get('sl')}, Target {pe_sig.get('target')} | ✅ Order placed")
                            
                            # Check signals for low strike (only if high strike didn't have signals)
                            if not has_ce_signal and not has_pe_signal:
                                low_signals = self.check_entry_signal_live(
                                    low_strike.get('ce_token'),
                                    low_strike.get('pe_token'),
                                    low_strike.get('ce_high'),
                                    low_strike.get('pe_high')
                                )
                                
                                if low_signals.get('success'):
                                    ce_sig = low_signals.get('ce_signal', {})
                                    pe_sig = low_signals.get('pe_signal', {})
                                    
                                    # Log what was received
                                    logger.info(f"LOW STRIKE Signals - CE has_signal: {ce_sig.get('has_signal')}, PE has_signal: {pe_sig.get('has_signal')}")
                                    if not ce_sig.get('has_signal'):
                                        logger.info(f"CE No Signal Reason: {ce_sig.get('reason', 'Unknown')}")
                                    if not pe_sig.get('has_signal'):
                                        logger.info(f"PE No Signal Reason: {pe_sig.get('reason', 'Unknown')}")
                                    
                                    # Update Excel logging data if signals exist
                                    if ce_sig.get('has_signal'):
                                        has_ce_signal = True
                                        ce_entry_price = ce_sig.get('entry_price')
                                        ce_sl = ce_sig.get('sl')
                                        ce_target = ce_sig.get('target')
                                        ce_high_val = low_strike.get('ce_high')
                                    
                                    if pe_sig.get('has_signal'):
                                        has_pe_signal = True
                                        pe_entry_price = pe_sig.get('entry_price')
                                        pe_sl = pe_sig.get('sl')
                                        pe_target = pe_sig.get('target')
                                        pe_high_val = low_strike.get('pe_high')
                                    
                                    # Only update trades if there are actual signals
                                    if ce_sig.get('has_signal') or pe_sig.get('has_signal'):
                                        self.update_active_trades(low_signals, strike_data=low_strike)
                                        self.log_signal(low_signals)
                                    
                                    if ce_sig.get('has_signal'):
                                        logger.info(f"📊 LOW STRIKE CE SIGNAL: Entry {ce_sig.get('entry_price')}, SL {ce_sig.get('sl')}, Target {ce_sig.get('target')} | ✅ Order placed")
                                    
                                    if pe_sig.get('has_signal'):
                                        logger.info(f"📊 LOW STRIKE PE SIGNAL: Entry {pe_sig.get('entry_price')}, SL {pe_sig.get('sl')}, Target {pe_sig.get('target')} | ✅ Order placed")
                        
                        # LOG EVERY 5-MINUTE CHECK TO EXCEL (both High + Low strike)
                        excel_logger.log_signal_check(
                            timestamp=check_timestamp,
                            ce_prev_high=ce_high_val,
                            ce_prev_low=None,  # Not tracked currently
                            pe_prev_high=pe_high_val,
                            pe_prev_low=None,  # Not tracked currently
                            ce_signal=has_ce_signal,
                            pe_signal=has_pe_signal,
                            ce_entry_price=ce_entry_price,
                            pe_entry_price=pe_entry_price,
                            ce_sl=ce_sl,
                            pe_sl=pe_sl,
                            ce_target=ce_target,
                            pe_target=pe_target,
                            notes="High Strike Check" if high_strike.get('success') else "High Strike data unavailable"
                        )

                        # Low strike log (even if no signals)
                        low_ce_high = low_strike.get('ce_high')
                        low_pe_high = low_strike.get('pe_high')
                        excel_logger.log_signal_check(
                            timestamp=check_timestamp,
                            ce_prev_high=low_ce_high,
                            ce_prev_low=None,  # Not tracked currently
                            pe_prev_high=low_pe_high,
                            pe_prev_low=None,  # Not tracked currently
                            ce_signal=has_ce_signal,
                            pe_signal=has_pe_signal,
                            ce_entry_price=ce_entry_price,
                            pe_entry_price=pe_entry_price,
                            ce_sl=ce_sl,
                            pe_sl=pe_sl,
                            ce_target=ce_target,
                            pe_target=pe_target,
                            notes="Low Strike Check" if low_strike.get('success') else "Low Strike data unavailable"
                        )
                        
                    else:
                        logger.warning(f"Failed to fetch live data: {live_data.get('error')}")
                        
                        # Log failed check to Excel
                        excel_logger.log_signal_check(
                            timestamp=check_timestamp,
                            notes=f"Failed to fetch data: {live_data.get('error', 'Unknown error')}"
                        )
                
                # Sleep 1 second and check again
                time_module.sleep(1)
            
            except Exception as e:
                logger.error(f"Error in ENTRY monitoring loop: {str(e)}", exc_info=True)
                time_module.sleep(1)

    def _sl_monitor_loop(self) -> None:
        """
        SL/Target monitoring loop - runs in background thread.
        Checked every 3 seconds.
        """
        logger.info(f"Starting SL/TARGET signal monitoring for {self.symbol} [Instance ID: {id(self)}]")

        
        while self.is_sl_monitoring:
            try:
                # Check if within market hours
                if not self.is_market_hours():
                    if self.is_market_day():
                        logger.debug("Outside market hours, waiting...")
                    time_module.sleep(1)
                    continue
                
                # ====== SL/TARGET CHECK (3-second intervals) ======
                if self.should_check_sl_target_now():
                    check_time = self.last_sl_target_check_time
                    if check_time:
                        logger.debug(f"[3-sec Check] Monitoring SL/Target for {self.symbol} at {check_time.strftime('%H:%M:%S.%f')[:-3]}")
                    else:
                        logger.debug(f"[3-sec Check] Monitoring SL/Target for {self.symbol}")
                    
                    # Check active trades for SL/Target hits
                    if self.active_trades:
                        logger.debug(f"Active trades to monitor: {list(self.active_trades.keys())}")
                        self.check_sl_target_for_active_trades(check_timestamp=check_time)
                    else:
                        logger.debug("No active trades to monitor")
                
                # ====== MARKET CLOSE CHECK (3:20 PM) ======
                # Force exit only trades that are STILL ACTIVE on the broker
                # This prevents trying to close trades already closed by SL hits
                current_time = datetime.now().time()
                if current_time >= time(15, 20, 0) and not self.market_close_processed:  # 3:20 PM IST
                    logger.info(f"🔴 Market close time (3:20 PM) reached - Checking for active trades to square off")
                    
                    if self.active_trades:
                        logger.info(f"Found {len(self.active_trades)} active trades to potentially close")
                        market_close_timestamp = datetime.now()
                        
                        # Get active orders from broker to sync state
                        active_orders = []
                        try:
                            if self.live_trading:
                                active_orders = self.kite.orders() if self.kite else []
                                logger.info(f"Fetched {len(active_orders)} active orders from broker")
                            else:
                                logger.info("Demo mode - skipping broker order fetch")
                        except Exception as e:
                            logger.warning(f"Could not fetch active orders from broker: {e}")
                            active_orders = []
                        
                        # Create set of active order IDs for quick lookup
                        active_order_ids = {str(order.get('order_id')) for order in active_orders if order.get('order_id')}
                        logger.info(f"📋 Active order IDs on broker: {active_order_ids if active_order_ids else 'None'}")
                        
                        closed_trades = []
                        skipped_trades = []
                        
                        with self.active_trades_lock:
                            for side in list(self.active_trades.keys()):
                                trade = self.active_trades.get(side)
                                if not trade:
                                    continue
                                
                                logger.info(f"Processing {side} trade: Status={trade.get('status')}, OrderID={trade.get('order_id')}")
                                
                                if trade.get('status') == 'OPEN':
                                    entry_order_id = trade.get('order_id')
                                    
                                    # Check if this order was placed by THIS SYSTEM
                                    is_system_order = entry_order_id is not None
                                    
                                    # Check if entry order was filled (COMPLETE) on broker
                                    # A filled BUY order means we have a position to close
                                    is_position_open = False
                                    if is_system_order:
                                        for order in active_orders:
                                            if str(order.get('order_id')) == str(entry_order_id):
                                                order_status = order.get('status', '').upper()
                                                # COMPLETE = filled, position exists
                                                # TRIGGER PENDING = SL order waiting (our SL not yet hit)
                                                if order_status == 'COMPLETE':
                                                    is_position_open = True
                                                logger.info(f"{side} Broker order {entry_order_id}: status={order_status}")
                                                break
                                        else:
                                            # Order not found in broker orders at all
                                            # Assume position exists if we have the trade in our system
                                            logger.warning(f"{side} Order {entry_order_id} not found in broker orders - assuming position open")
                                            is_position_open = True
                                    
                                    logger.info(f"{side} Trade - System Order: {is_system_order}, Position Open: {is_position_open}")
                                    
                                    if is_system_order and is_position_open:
                                        # This is our order and it's still active - close it
                                        token = trade.get('token')
                                        current_price = None
                                        
                                        try:
                                            if token:
                                                current_price = self.get_current_price(token)
                                                logger.info(f"{side} Current price from token {token}: {current_price}")
                                            else:
                                                logger.warning(f"{side} No token available, using entry price")
                                        except Exception as e:
                                            logger.warning(f"Error fetching current price for {side}: {e}")
                                        
                                        if not current_price:
                                            current_price = trade.get('entry_price')
                                        
                                        entry_price = trade.get('entry_price', 0)
                                        strike = trade.get('strike')
                                        pnl = current_price - entry_price if current_price else 0
                                        
                                        logger.info(f"🔴 {side} Market Close Square Off: Entry {entry_price:.2f}, Exit {current_price:.2f}, P&L: {pnl:+.2f} | Order ID: {entry_order_id}")
                                        
                                        # Place SELL order to close the position
                                        try:
                                            if strike and current_price:
                                                sell_order_id = self.place_sell_order(
                                                    side=side,
                                                    strike=strike,
                                                    exit_price=current_price,
                                                    exit_reason="Market Close (3:20 PM)"
                                                )
                                                logger.info(f"✅ {side} Square off order placed at market close | Exit Order ID: {sell_order_id if sell_order_id else 'N/A'}")
                                                closed_trades.append(side)
                                            else:
                                                logger.error(f"❌ Cannot place sell order for {side}: Strike={strike}, Price={current_price}")
                                                skipped_trades.append(f"{side} (missing strike/price)")
                                        except Exception as e:
                                            logger.error(f"Error placing sell order for {side} at market close: {e}", exc_info=True)
                                            skipped_trades.append(f"{side} (order error: {str(e)})")
                                        
                                        # Log to Excel sheets
                                        try:
                                            entry_time = trade.get('entry_time', 'N/A')
                                            entry_order_id = trade.get('order_id', 'N/A')
                                            
                                            excel_logger.log_sl_target_check(
                                                timestamp=market_close_timestamp,
                                                side=side,
                                                strike=strike,
                                                current_price=current_price if current_price else entry_price,
                                                entry_price=entry_price,
                                                initial_sl=trade.get('sl'),
                                                target=trade.get('target'),
                                                target_hit=trade.get('target_hit'),
                                                check_reason="MARKET_CLOSE"
                                            )
                                            
                                            excel_logger.log_trade(
                                                order_type='SELL',
                                                option_type=side,
                                                strike=strike,
                                                entry_price=entry_price,
                                                current_price=current_price if current_price else entry_price,
                                                target=trade.get('target'),
                                                stop_loss=trade.get('sl'),
                                                pnl=pnl,
                                                status='MARKET_CLOSE',
                                                notes=f"Market Close Square-Off (3:20 PM) | Entry Time: {entry_time} | Entry Order ID: {entry_order_id} | Entry Price: {entry_price:.2f} | Stop Loss: {trade.get('sl', 'N/A')} | Exit Price: {current_price:.2f} | P&L: {pnl:+.2f}"
                                            )
                                            logger.info(f"✅ {side} trade logged to Excel with entry and stop loss details")
                                        except Exception as e:
                                            logger.warning(f"Error logging trade to Excel for {side}: {e}")
                                        
                                        # Close the trade in system
                                        try:
                                            self.close_trade(side, current_price if current_price else entry_price, "Market Close (3:20 PM)")
                                        except Exception as e:
                                            logger.error(f"Error closing trade {side} in system: {e}")
                                    
                                    elif is_system_order and not is_position_open:
                                        # Our order but already closed (likely by SL hit) - just mark as closed in system
                                        logger.info(f"⏹️  {side} Order {entry_order_id} already closed on broker (likely by SL hit) - marking as closed")
                                        try:
                                            entry_time = trade.get('entry_time', 'N/A')
                                            entry_price = trade.get('entry_price', 0)
                                            exit_price = trade.get('exit_price', entry_price)
                                            pnl = exit_price - entry_price if entry_price else 0
                                            
                                            # Log the already-closed trade
                                            excel_logger.log_trade(
                                                order_type='SELL',
                                                option_type=side,
                                                strike=trade.get('strike'),
                                                entry_price=entry_price,
                                                current_price=exit_price if exit_price else entry_price,
                                                target=trade.get('target'),
                                                stop_loss=trade.get('sl'),
                                                pnl=pnl,
                                                status='SL_HIT',
                                                notes=f"Trade closed by Stop Loss Hit (checked at 3:20 PM) | Entry Time: {entry_time} | Entry Order ID: {entry_order_id} | Entry Price: {entry_price:.2f} | Stop Loss: {trade.get('sl', 'N/A')} | Exit Price: {exit_price:.2f} | P&L: {pnl:+.2f}"
                                            )
                                            logger.info(f"✅ {side} already-closed trade logged to Excel")
                                            
                                            self.close_trade(side, trade.get('entry_price'), "Already closed on broker (SL hit)")
                                            closed_trades.append(f"{side} (already closed by SL)")
                                        except Exception as e:
                                            logger.error(f"Error closing already-closed trade {side}: {e}")
                                    
                                    else:
                                        # Not our order - don't touch it (manual trades or from other sources)
                                        logger.info(f"⏭️  {side} Order is not from this system - skipping (allowing manual trades to remain open)")
                                        skipped_trades.append(f"{side} (not system order)")
                                else:
                                    logger.info(f"{side} Trade is not OPEN (Status: {trade.get('status')}) - skipping")
                                    skipped_trades.append(f"{side} (not open)")
                        
                        # Summary logging
                        logger.info(f"🏁 Market close square off complete:")
                        logger.info(f"   ✅ Closed: {closed_trades if closed_trades else 'None'}")
                        logger.info(f"   ⏭️  Skipped: {skipped_trades if skipped_trades else 'None'}")
                        
                        # Mark that market close has been processed
                        self.market_close_processed = True
                    else:
                        logger.info("No active trades at market close (3:20 PM)")
                        self.market_close_processed = True
                
                # Sleep 1 second and check again
                time_module.sleep(1)
                
            except Exception as e:
                logger.error(f"Error in SL monitoring loop: {str(e)}", exc_info=True)
                time_module.sleep(1)
    
    def start_monitoring(self, monitor_entries: bool = True, monitor_sl: bool = True) -> Dict[str, bool]:
        """
        Start live signal monitoring in background threads.
        
        Args:
            monitor_entries: Whether to start entry monitoring loop
            monitor_sl: Whether to start SL/Target monitoring loop
            
        Returns:
            Dictionary with status of each monitor
        """
        result = {'entries_started': False, 'sl_started': False}
        
        if not self.is_market_day():
            logger.warning("Not a market trading day")
            return result
        
        # Acquire process lock before starting threads
        if not self._acquire_process_lock():
            logger.error(f"❌ Failed to start monitoring: Another instance is already running for {self.username}")
            return result

        
        # Start Entry Monitoring
        if monitor_entries:
            if self.is_entry_monitoring:
                logger.warning("Entry Monitoring already running")
            else:
                self.is_entry_monitoring = True
                self.entry_monitor_thread = threading.Thread(
                    target=self._entry_monitor_loop,
                    name=f"Intraday920EntryMonitor-{self.symbol}",
                    daemon=True
                )
                self.entry_monitor_thread.start()
                logger.info(f"Live ENTRY monitoring started for {self.symbol} [Instance ID: {id(self)}]")
                result['entries_started'] = True

        # Start SL Monitoring
        if monitor_sl:
            if self.is_sl_monitoring:
                logger.warning("SL Monitoring already running")
            else:
                self.is_sl_monitoring = True
                self.sl_monitor_thread = threading.Thread(
                    target=self._sl_monitor_loop,
                    name=f"Intraday920SLMonitor-{self.symbol}",
                    daemon=True
                )
                self.sl_monitor_thread.start()
                logger.info(f"Live SL/TARGET monitoring started for {self.symbol} [Instance ID: {id(self)}]")
                result['sl_started'] = True
        
        self.is_monitoring = self.is_entry_monitoring or self.is_sl_monitoring
        return result
    
    def stop_monitoring(self, stop_entries: bool = True, stop_sl: bool = True) -> Dict[str, bool]:
        """
        Stop live signal monitoring.
        
        Args:
            stop_entries: Whether to stop entry monitoring
            stop_sl: Whether to stop SL monitoring
            
        Returns:
            Dictionary with status of each stop action
        """
        result = {'entries_stopped': False, 'sl_stopped': False}
        
        if stop_entries and self.is_entry_monitoring:
            self.is_entry_monitoring = False
            if self.entry_monitor_thread:
                self.entry_monitor_thread.join(timeout=5)
            logger.info(f"Live ENTRY monitoring stopped for {self.symbol}")
            result['entries_stopped'] = True
            
        if stop_sl and self.is_sl_monitoring:
            self.is_sl_monitoring = False
            if self.sl_monitor_thread:
                self.sl_monitor_thread.join(timeout=5)
            logger.info(f"Live SL/TARGET monitoring stopped for {self.symbol}")
            result['sl_stopped'] = True
            
        # Release lock if everything stopped
        if not self.is_entry_monitoring and not self.is_sl_monitoring:
            self._release_process_lock()
            
        self.is_monitoring = self.is_entry_monitoring or self.is_sl_monitoring
        return result
    
    def get_active_trades(self) -> Dict[str, Any]:
        """
        Get current active trades.
        
        Returns:
            Dictionary of active trades
        """
        return self.active_trades.copy()
    
    def get_today_signals(self) -> List[Dict[str, Any]]:
        """
        Get all signals generated today.
        
        Returns:
            List of today's signals
        """
        return self.today_signals.copy()
    
    def close_trade(self, side: str, exit_price: float, exit_reason: str = "Manual") -> Dict[str, Any]:
        """
        Close an active trade.
        
        Args:
            side: 'CE' or 'PE'
            exit_price: Exit price
            exit_reason: Reason for exit
            
        Returns:
            Trade summary
        """
        with self.active_trades_lock:  # RLock allows nested acquisition from check_sl_target
            if side not in self.active_trades:
                return {'success': False, 'error': f'No active {side} trade'}
            
            trade = self.active_trades[side]
            entry_price = trade.get('entry_price', 0)
            pnl = exit_price - entry_price
            pnl_pct = (pnl / entry_price * 100) if entry_price else 0
            
            trade['exit_price'] = exit_price
            trade['exit_reason'] = exit_reason
            trade['exit_time'] = datetime.now().isoformat()
            trade['pnl'] = round(pnl, 2)
            trade['pnl_pct'] = round(pnl_pct, 2)
            trade['status'] = 'CLOSED'
            
            logger.info(f"🔴 {side} trade closed: Entry {entry_price}, Exit {exit_price}, PnL {pnl} ({pnl_pct}%)")
            
            return {
                'success': True,
                'trade': trade
            }
    
    def reset_daily(self) -> None:
        """Reset daily signals and trades (call at market open)."""
        self.active_trades = {}
        self.today_signals = []
        self.last_entry_check_time = None
        self.reset_daily_entries()  # Reset daily entry tracking
        logger.info(f"Daily monitoring reset for {self.symbol}")
